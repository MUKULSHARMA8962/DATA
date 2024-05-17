import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from lnd_file import dfs

for key, value in dfs.items():
    # print(f"{key} - Benefit Package DataFrame:")
    BenefitPackage = value['Benefit Package']
    # - Benefit Information DataFrame
    BenefitInformation = value['Benefit Information']
    
    first_row_value = BenefitPackage.iloc[0]
    hios_plan_id = first_row_value['HIOS Plan ID*(Standard Component)']
    if pd.notna(hios_plan_id):
        ke = hios_plan_id[:5]  # '32753'
        state = hios_plan_id[5:7]  # 'MO'


      


# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Set the title of the sheet and freeze panes
ws.title = "2025 Plans & Benefits Template v14.0"
ws.freeze_panes = 'A8'  # Freeze the rows above A8

# Define the borders and fill colors
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green color for headers
instruction_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green color for instructions

# Set initial instructional content with colors
instructions = [
    ("2025 Plans & Benefits Template v14.0", "A1:C1"),
    ("HIOS Issuer ID*: 32753", "A2:A2"),
    ("Issuer State*: MO", "A3:A3"),
    ("Market Coverage*: Individual", "A4:A4"),
    ("Dental Only Plan*: No", "A5:A5"),
    ("To use this template, please review the user guide and instructions. All fields with an asterisk (*) are required", "G1:K1"),
    ("you will need to save the latest version of the add-in file (PlansBenefitsAddInPY25.xlam) on your machine.", "G2:K2"),
    ("To create the cost share variance worksheet and enter the cost sharing amounts for both individual and SHOP (small group) markets, use the Create Cost Share Variances macro.", "G3:K3"),
    ("To create additional Benefits Package worksheets, use the Create New Benefits Package macro.", "G4:K4"),
    ("o populate the benefits on the Benefits Package worksheet with your State EHB Standards, use the Refresh EHB macro.", "G5:K5")
]


for text, cell_range in instructions:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(':')[0]] 
    cell.value = text
    if cell_range == "A1:E1":  # Check if it's the title cell range
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
     # Apply light green fill for instructional content
    cell.fill = instruction_fill if "No" in text else header_fill


# Fill in the headers
for text, cell_range in instructions:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(':')[0]]
    cell.value = text
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.fill = instruction_fill

# Sample data that corresponds to the headers
data = {
    "B2": "32753",
    "B3": "MO",
    "B4": "Individual",
    "B5": "No"
}

# Populate data into Column B
for cell_address, value in data.items():
    cell = ws[cell_address]
    cell.value = value
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = thin_border


# Set up headers in row 7 with coloring
headers_row_7 = [
    "HIOS Plan ID*\n(Standard Component)", "Plan Marketing Name*", "HIOS Product ID*",
    "Network ID*", "Service Area ID*", "Formulary ID*", "New/Existing Plan?*", "Plan Type*",
    "Level of Coverage*", "Design Type*", "Unique Plan Design?*", "QHP/Non-QHP*",
    "Notice Required\nfor Pregnancy*", "Plan Level Exclusions",
    "Limited Cost Sharing Plan Variation - Est Advanced Payment",
    "Does this plan offer Composite Rating?*", "Child-Only Offering*", "Child Only Plan ID",
    "Tobacco Wellness Program Offered*", "Disease Management Programs Offered",
    "EHB Percent of Total Premium", "EHB Apportionment for Pediatric Dental",
    "Guaranteed Rate", "Plan Effective Date*", "Plan Expiration Date",
    "Out of Country Coverage*", "Out of Country Coverage Description",
    "Out of Service Area Coverage*", "Out of Service Area Coverage Description",
    "National Network*"
]
# Headers for row 20
headers_row_20 = [
 "Out of Country Coverage Description",
    "Out of Service Area Coverage*", 
    "Out of Service Area Coverage Description",
    "National Network*", 
     "Benefits", 
     "EHB	Is this Benefit Covered?",
     "Quantitative Limit on Service", 
     "Limit Quantity", 
     "Limit Unit", 
     "Exclusions", 
     "Benefit Explanation", 
     "Explanation",
]
# Function to set headers
def set_headers(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = header_fill

# Apply headers on row 7
set_headers(ws, 7, headers_row_7)

# Apply headers on row 20
set_headers(ws, 20, headers_row_20)

# Set column widths for visibility
column_widths = [30, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 30, 30, 30, 30]
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Sample data to be inserted
data = {
    'HIOS Plan ID': ['101', '102'],
    'Plan Marketing Name': ['Silver Plan', 'Gold Plan'],
    # Extend this dictionary to include all necessary data for each header
}

# Create DataFrame
df = pd.DataFrame(data)

# Function to populate data from DataFrame
def populate_data_from_df(ws, start_row, df):
    for index, row in df.iterrows():
        for col_index, (key, value) in enumerate(row.items()):
            cell = ws.cell(row=start_row + index, column=col_index + 1)
            cell.value = value
            cell.border = thin_border

# Populate the data from the data frame into the worksheet
populate_data_from_df(ws, start_row=8, df=df)  # Assuming data starts at row 8 below the headers

# Create Sheet 2 and apply formatting
ws2 = wb.create_sheet(title="Cost Share Variances ")

# Define fills for different types of cells
instruction_fill = PatternFill(start_color="CCCC99", end_color="CCCC99", fill_type="solid")  # Light green fill
header_fill = PatternFill(start_color="CCCC99", end_color="CCCC99", fill_type="solid")  # Light orange fill, for headers

# Set initial instructional content with colors
instructions = [
    ("All Fields with an asterisk(*) are required", "A1:C1"),
    ("", "A2:A2"),
]
# Function to apply instructions
def apply_instructions(sheet):
    for text, cell_range in instructions:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(':')[0]]  # Get the first cell of the range
        cell.value = text
        cell.alignment = Alignment(horizontal='center' if "A1:C1" in cell_range else 'left', vertical='center', wrap_text=True)
        # Apply light green fill for instructional content, but apply a different fill for the title
        cell.fill = header_fill if "A1:C1" in cell_range else instruction_fill

apply_instructions(ws2)
headers_sheet2 = [
    "Header 1", "Header 2", "Header 3", "Header 4", "Header 5"
]

# Apply additional headers to Sheet2 starting from the fourth column of the third row
for col, header in enumerate(headers_sheet2[1:], start=1):
    cell = ws2.cell(row=3, column=col)
    cell.value = header
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True, color='FFFFFF')
    cell.border = thin_border
    cell.fill = PatternFill(start_color="333399", end_color="333399", fill_type="solid")  # Dark blue fill


# Add dummy data under headers in Sheet 2
dummy_data = [
    ["Data 1-1", "Data 1-2", "Data 1-3", "Data 1-4", "Data 1-5"],
    ["Data 2-1", "Data 2-2", "Data 2-3", "Data 2-4", "Data 2-5"],
    ["Data 3-1", "Data 3-2", "Data 3-3", "Data 3-4", "Data 3-5"],
    ["Data 4-1", "Data 4-2", "Data 4-3", "Data 4-4", "Data 4-5"]
]

# Populate dummy data starting from row 4
for row_index, row_data in enumerate(dummy_data, start=4):  # Start from row 4
    for col_index, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_index, column=col_index)  # Use row_index which now starts at 4
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = instruction_fill  # Using the same light green color for consistency

# Save the workbook
wb.save('Custom_Structured_Excel_File_v2.xlsx')
