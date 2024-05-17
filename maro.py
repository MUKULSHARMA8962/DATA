import pandas as pd
import openpyxl
from lnd_file import dfs

# Load the workbook with VBA macros enabled
file_path = 'PY2025PlansBenefitsTemplate (1) .xlsm'
workbook = openpyxl.load_workbook(file_path, keep_vba=True)

# Load the CSV file into a DataFrame
filename = "qhp_midsection (3) (1).csv"
mindsection_df = pd.read_csv(filename)

# Template sheet name for copying the structure
template_sheet_name = 'Benefits Package 2'

# Check if the template sheet exists in the workbook
if template_sheet_name not in workbook.sheetnames:
    raise ValueError(f"Template sheet '{template_sheet_name}' not found in the workbook.")

# Map keys to the expected sheet names
key_mapping = {f'Benefit PKG {i}': f'Benefits Package {i}' for i in range(1, len(dfs) + 1)}

# Print the key mapping for debugging
print(f"Key Mapping: {key_mapping}")

# Flag to track if any data was written
data_written = False

# Iterate over each key, value pair in dfs
for key, value in dfs.items():
    # Map the key to the expected sheet name
    if key in key_mapping:
        sheet_name = key_mapping[key]
        print(f"Processing key: {key}, mapped to sheet: {sheet_name}")
        
        # Check if the sheet exists in the workbook
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            # Create a new sheet with the same structure as the template
            template_sheet = workbook[template_sheet_name]
            sheet = workbook.copy_worksheet(template_sheet)
            sheet.title = sheet_name
        
        # Extract the required data
        BenefitPackage = value['Benefit Package']
        
        # Perform the left join with an indicator column
        left_join = pd.merge(
            mindsection_df,
            BenefitPackage.rename(columns={'HIOS Plan ID*(Standard Component)': 'HIOS Plan ID* (Standard Component)'}),
            on='HIOS Plan ID* (Standard Component)',
            how='left',
            indicator=True
        )
        
        # Filter the rows where the match happened (both columns match)
        filtered_left_join = left_join[left_join['_merge'] == 'both'].drop(columns='_merge')
        
        # Drop additional columns as needed
        columns_to_drop = ['lastModificationDate', 'Refresh Date', 'Plan Marketing Name*_x', 'Level of Coverage*_x', 'Plan Type*_x']
        filtered_left_join = filtered_left_join.drop(columns=columns_to_drop)

        # Rename columns as needed
        columns_to_rename = {'Plan Marketing Name*_y': 'Plan Marketing Name*', 'Level of Coverage*_y': 'Level of Coverage*', 'Plan Type*_y': 'Plan Type*'}
        filtered_left_join = filtered_left_join.rename(columns=columns_to_rename)

        # Ensure there are no duplicate rows
        filtered_left_join = filtered_left_join.drop_duplicates()

        BenefitInformation = value['Benefit Information']
        first_row_value = BenefitPackage.loc[0]
        hios_plan_id = first_row_value['HIOS Plan ID*(Standard Component)']
        
        if pd.notna(hios_plan_id):
            ke = hios_plan_id[:5]  # Extract HIOS Plan ID
            state = hios_plan_id[5:7]  # Extract State


        benefit_data = BenefitInformation.head(20)

        # Function to populate data from DataFrame
        def populate_data_from_df(ws, start_row, df):
            for index, row in df.iterrows():
                for col_index, (key, value) in enumerate(row.items()):
                    cell = ws.cell(row=start_row + index, column=col_index + 1)
                    cell.value = value

            populate_data_from_df(ws, start_row=21, df=benefit_data)  # Assuming data starts at row 21 below the headers in row 20









        # Write DataFrame to the sheet starting from row 8
        start_row = 8

        # Iterate over the DataFrame and write values to the corresponding cells
        for row_idx, row in filtered_left_join.iterrows():
            for col_idx, value in enumerate(row):
                cell = sheet.cell(row=start_row + row_idx, column=col_idx + 1)
                cell.value = value

        # Set the flag to True if data was written
        data_written = True
        print(f"Data successfully written to sheet '{sheet_name}' for key '{key}'")
        # Remove extra empty rows after writing the data
        max_row = sheet.max_row
        while max_row > start_row:
            row_values = [cell.value for cell in sheet[max_row]]
            if all(value is None for value in row_values):
                sheet.delete_rows(max_row)
            else:
                break
            max_row -= 1
        

# If no data was written, print a message
if not data_written:
    print(f"No data was written to any sheet as no matching key was found in 'dfs'.")

# Save the workbook with a new name
new_file_path = 'Modified_PY2025PlansBenefitsTemplate.xlsm'
workbook.save(new_file_path)
print(f"Workbook saved as '{new_file_path}'")
