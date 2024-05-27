import pandas as pd
import openpyxl
from openpyxl.styles import Font
from lnd_file import dfs
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
# Load the workbook with VBA macros enabled
file_path = 'base_file//PY2025PlansBenefitsTemplate.xlsm'
workbook = openpyxl.load_workbook(file_path, keep_vba=True)

# Load the CSV files into DataFrames
mindsection_df = pd.read_csv("input_file//qhp_midsection.csv")
hp_a2ch_df = pd.read_csv("input_file//qhp_a2ch.csv")
# Step 2: List of columns to delete
columns_to_delete = [
    'Plan Modified date', 'Refresh Date','Plan Marketing Name_old*','URL for Summary of Benefits & Coverage', 'Plan Brochure']
hp_a2ch_df.drop(columns=columns_to_delete, inplace=True)

qhp_costshare_df = pd.read_csv("input_file//qhp_costshare.csv")
column_mapping = {
    'hios_plan_id_with_variant': '0HIOS_Plan_ID_With_Variant',
    'Benefit1_INN_Copay': '2_Primary Care Visit to Treat an Injury or Illness_Benefit1_INN_Copay',
    'Benefit1_INN2_Copay': '2_Primary Care Visit to Treat an Injury or Illness_Benefit1_INN2_Copay',
    'Benefit1_OON_Copay': '2_Primary Care Visit to Treat an Injury or Illness_Benefit1_OON_Copay',
    'benefit1_inn_coinsurance': '2_Primary Care Visit to Treat an Injury or Illness_Benefit1_INN_Coinsurance',
    'benefit1_inn2_coinsurance': '2_Primary Care Visit to Treat an Injury or Illness_Benefit1_INN2_Coinsurance',
    'benefit1_oon_coinsurance': '2_Primary Care Visit to Treat an Injury or Illness_Benefit1_OON_Coinsurance',
    'Benefit2_INN_Copay': '12_Specialist Visit_Benefit2_INN_Copay',
    'Benefit2_INN2_Copay': '12_Specialist Visit_Benefit2_INN2_Copay',
    'Benefit2_OON_Copay': '12_Specialist Visit_Benefit2_OON_Copay',
    'Benefit2_INN_Coinsurance': '12_Specialist Visit_Benefit2_INN_Coinsurance',
    'Benefit2_INN2_Coinsurance': '12_Specialist Visit_Benefit2_INN2_Coinsurance',
    'Benefit2_OON_Coinsurance': '12_Specialist Visit_Benefit2_OON_Coinsurance',
    'Benefit3_INN_Copay': '2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_INN_Copay',
    'Benefit3_INN2_Copay': '2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_INN2_Copay',
    'Benefit3_OON_Copay': '2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_OON_Copay',
    'Benefit3_INN_Coinsurance': '2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_INN_Coinsurance',
    'Benefit3_INN2_Coinsurance': '2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_INN2_Coinsurance',
    'Benefit3_OON_Coinsurance': '2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_OON_Coinsurance',
    'Benefit4_INN_Copay': '47_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_INN_Copay',
    'Benefit4_INN2_Copay': '48_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_INN2_Copay',
    'Benefit4_OON_Copay': '47_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_OON_Copay',
    'Benefit4_INN_Coinsurance': '47_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_INN_Coinsurance',
    'Benefit4_INN2_Coinsurance': '48_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_INN2_Coinsurance',
    'Benefit4_OON_Coinsurance': '47_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_OON_Coinsurance',
    'Benefit5_INN_Copay': '50_Outpatient Surgery Physician/Surgical Services_Benefit5_INN_Copay',
    'Benefit5_INN2_Copay': '50_Outpatient Surgery Physician/Surgical Services_Benefit5_INN2_Copay',
    'Benefit5_OON_Copay': '50_Outpatient Surgery Physician/Surgical Services_Benefit5_OON_Copay',
    'Benefit5_INN_Coinsurance': '50_Outpatient Surgery Physician/Surgical Services_Benefit5_INN_Coinsurance',
    'Benefit5_INN2_Coinsurance': '50_Outpatient Surgery Physician/Surgical Services_Benefit5_INN2_Coinsurance',
    'Benefit5_OON_Coinsurance': '50_Outpatient Surgery Physician/Surgical Services_Benefit5_OON_Coinsurance',
    'Benefit6_INN_Copay': '254_Hospice Services_Benefit6_INN_Copay',
    'Benefit6_INN2_Copay': '254_Hospice Services_Benefit6_INN2_Copay',
    'Benefit6_OON_Copay': '254_Hospice Services_Benefit6_OON_Copay',
    'Benefit6_INN_Coinsurance': '254_Hospice Services_Benefit6_INN_Coinsurance',
    'Benefit6_INN2_Coinsurance': '254_Hospice Services_Benefit6_INN2_Coinsurance',
    'Benefit6_OON_Coinsurance': '254_Hospice Services_Benefit6_OON_Coinsurance',
    'Benefit7_INN_Copay': '253_Private-Duty Nursing_Benefit7_INN_Copay',
    'Benefit7_INN2_Copay': '253_Private-Duty Nursing_Benefit7_INN2_Copay',
    'Benefit7_OON_Copay': '253_Private-Duty Nursing_Benefit7_OON_Copay',
    'Benefit7_INN_Coinsurance': '253_Private-Duty Nursing_Benefit7_INN_Coinsurance',
    'Benefit7_INN2_Coinsurance': '253_Private-Duty Nursing_Benefit7_INN2_Coinsurance',
    'Benefit7_OON_Coinsurance': '253_Private-Duty Nursing_Benefit7_OON_Coinsurance',
    'Benefit8_INN_Copay': '33_Urgent Care Centers or Facilities_Benefit8_INN_Copay',
    'Benefit8_INN2_Copay': '33_Urgent Care Centers or Facilities_Benefit8_INN2_Copay',
    'Benefit8_OON_Copay': '33_Urgent Care Centers or Facilities_Benefit8_OON_Copay',
    'Benefit8_INN_Coinsurance': '33_Urgent Care Centers or Facilities_Benefit8_INN_Coinsurance',
    'Benefit8_INN2_Coinsurance': '33_Urgent Care Centers or Facilities_Benefit8_INN2_Coinsurance',
    'Benefit8_OON_Coinsurance': '33_Urgent Care Centers or Facilities_Benefit8_OON_Coinsurance',
    'Benefit9_INN_Copay': '98_Home Health Care Services_Benefit9_INN_Copay',
    'Benefit9_INN2_Copay': '98_Home Health Care Services_Benefit9_INN2_Copay',
    'Benefit9_OON_Copay': '98_Home Health Care Services_Benefit9_OON_Copay',
    'Benefit9_INN_Coinsurance': '98_Home Health Care Services_Benefit9_INN_Coinsurance',
    'Benefit9_INN2_Coinsurance': '98_Home Health Care Services_Benefit9_INN2_Coinsurance',
    'Benefit9_OON_Coinsurance': '98_Home Health Care Services_Benefit9_OON_Coinsurance',
    'Benefit10_INN_Copay': '43_Emergency Room Services_Benefit10_INN_Copay',
    'Benefit10_INN2_Copay': '43_Emergency Room Services_Benefit10_INN2_Copay',
    'Benefit10_OON_Copay': '43_Emergency Room Services_Benefit10_OON_Copay',
    'Benefit10_INN_Coinsurance': '43_Emergency Room Services_Benefit10_INN_Coinsurance',
    'Benefit10_INN2_Coinsurance': '43_Emergency Room Services_Benefit10_INN2_Coinsurance',
    'Benefit10_OON_Coinsurance': '43_Emergency Room Services_Benefit10_OON_Coinsurance',
    'Benefit11_INN_Copay': '103_Emergency Transportation/Ambulance_Benefit11_INN_Copay',
    'Benefit11_INN2_Copay': '103_Emergency Transportation/Ambulance_Benefit11_INN2_Copay',
    'Benefit11_OON_Copay': '103_Emergency Transportation/Ambulance_Benefit11_OON_Copay',
    'Benefit11_INN_Coinsurance': '103_Emergency Transportation/Ambulance_Benefit11_INN_Coinsurance',
    'Benefit11_INN2_Coinsurance': '103_Emergency Transportation/Ambulance_Benefit11_INN2_Coinsurance',
    'Benefit11_OON_Coinsurance': '103_Emergency Transportation/Ambulance_Benefit11_OON_Coinsurance',
    'Benefit12_INN_Copay': '79_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_INN_Copay',
    'Benefit12_INN2_Copay': '80_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_INN2_Copay',
    'Benefit12_OON_Copay': '79_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_OON_Copay',
    'Benefit12_INN_Coinsurance': '79_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_INN_Coinsurance',
    'Benefit12_INN2_Coinsurance': '80_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_INN2_Coinsurance',
    'Benefit12_OON_Coinsurance': '79_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_OON_Coinsurance',
    'Benefit13_INN_Copay': '93_Inpatient Physician and Surgical Services_Benefit13_INN_Copay',
    'Benefit13_INN2_Copay': '93_Inpatient Physician and Surgical Services_Benefit13_INN2_Copay',
    'Benefit13_OON_Copay': '93_Inpatient Physician and Surgical Services_Benefit13_OON_Copay',
    'Benefit13_INN_Coinsurance': '93_Inpatient Physician and Surgical Services_Benefit13_INN_Coinsurance',
    'Benefit13_INN2_Coinsurance': '93_Inpatient Physician and Surgical Services_Benefit13_INN2_Coinsurance',
    'Benefit13_OON_Coinsurance': '93_Inpatient Physician and Surgical Services_Benefit13_OON_Coinsurance',
    'Benefit14_INN_Copay': '81_Skilled Nursing Facility_Benefit14_INN_Copay',
    'Benefit14_INN2_Copay': '82_Skilled Nursing Facility_Benefit14_INN2_Copay',
    'Benefit14_OON_Copay': '81_Skilled Nursing Facility_Benefit14_OON_Copay',
    'Benefit14_INN_Coinsurance': '81_Skilled Nursing Facility_Benefit14_INN_Coinsurance',
    'Benefit14_INN2_Coinsurance': '82_Skilled Nursing Facility_Benefit14_INN2_Coinsurance',
    'Benefit14_OON_Coinsurance': '81_Skilled Nursing Facility_Benefit14_OON_Coinsurance',
    'Benefit15_INN_Copay': '6_Prenatal and Postnatal Care_Benefit15_INN_Copay',
    'Benefit15_INN2_Copay': '6_Prenatal and Postnatal Care_Benefit15_INN2_Copay',
    'Benefit15_OON_Copay': '6_Prenatal and Postnatal Care_Benefit15_OON_Copay',
    'Benefit15_INN_Coinsurance': '6_Prenatal and Postnatal Care_Benefit15_INN_Coinsurance',
    'Benefit15_INN2_Coinsurance': '6_Prenatal and Postnatal Care_Benefit15_INN2_Coinsurance',
    'Benefit15_OON_Coinsurance': '6_Prenatal and Postnatal Care_Benefit15_OON_Coinsurance',
    'Benefit16_INN_Copay': '79_Delivery and All Inpatient Services for Maternity Care_Benefit16_INN_Copay',
    'Benefit16_INN2_Copay': '80_Delivery and All Inpatient Services for Maternity Care_Benefit16_INN2_Copay',
    'Benefit16_OON_Copay': '79_Delivery and All Inpatient Services for Maternity Care_Benefit16_OON_Copay',
    'Benefit16_INN_Coinsurance': '79_Delivery and All Inpatient Services for Maternity Care_Benefit16_INN_Coinsurance',
    'Benefit16_INN2_Coinsurance': '80_Delivery and All Inpatient Services for Maternity Care_Benefit16_INN2_Coinsurance',
    'Benefit16_OON_Coinsurance': '79_Delivery and All Inpatient Services for Maternity Care_Benefit16_OON_Coinsurance',
    'Benefit17_INN_Copay': '246_Mental/Behavioral Health Outpatient Services_Benefit17_INN_Copay',
    'Benefit17_INN2_Copay': '246_Mental/Behavioral Health Outpatient Services_Benefit17_INN2_Copay',
    'Benefit17_OON_Copay': '246_Mental/Behavioral Health Outpatient Services_Benefit17_OON_Copay',
    'Benefit17_INN_Coinsurance': '246_Mental/Behavioral Health Outpatient Services_Benefit17_INN_Coinsurance',
    'Benefit17_INN2_Coinsurance': '246_Mental/Behavioral Health Outpatient Services_Benefit17_INN2_Coinsurance',
    'Benefit17_OON_Coinsurance': '246_Mental/Behavioral Health Outpatient Services_Benefit17_OON_Coinsurance',
    'Benefit18_INN_Copay': '83_Mental/Behavioral Health Inpatient Services_Benefit18_INN_Copay',
    'Benefit18_INN2_Copay': '84_Mental/Behavioral Health Inpatient Services_Benefit18_INN2_Copay',
    'Benefit18_OON_Copay': '83_Mental/Behavioral Health Inpatient Services_Benefit18_OON_Copay',
    'Benefit18_INN_Coinsurance': '83_Mental/Behavioral Health Inpatient Services_Benefit18_INN_Coinsurance',
    'Benefit18_INN2_Coinsurance': '84_Mental/Behavioral Health Inpatient Services_Benefit18_INN2_Coinsurance',
    'Benefit18_OON_Coinsurance': '83_Mental/Behavioral Health Inpatient Services_Benefit18_OON_Coinsurance',
    'Benefit19_INN_Copay': '246_Substance Abuse Disorder Outpatient Services_Benefit19_INN_Copay',
    'Benefit19_INN2_Copay': '246_Substance Abuse Disorder Outpatient Services_Benefit19_INN2_Copay',
    'Benefit19_OON_Copay': '246_Substance Abuse Disorder Outpatient Services_Benefit19_OON_Copay',
    'Benefit19_INN_Coinsurance': '246_Substance Abuse Disorder Outpatient Services_Benefit19_INN_Coinsurance',
    'Benefit19_INN2_Coinsurance': '246_Substance Abuse Disorder Outpatient Services_Benefit19_INN2_Coinsurance',
    'Benefit19_OON_Coinsurance': '246_Substance Abuse Disorder Outpatient Services_Benefit19_OON_Coinsurance',
    'Benefit20_INN_Copay': '83_Substance Abuse Disorder Inpatient Services_Benefit20_INN_Copay',
    'Benefit20_INN2_Copay': '84_Substance Abuse Disorder Inpatient Services_Benefit20_INN2_Copay',
    'Benefit20_OON_Copay': '83_Substance Abuse Disorder Inpatient Services_Benefit20_OON_Copay',
    'Benefit20_INN_Coinsurance': '83_Substance Abuse Disorder Inpatient Services_Benefit20_INN_Coinsurance',
    'Benefit20_INN2_Coinsurance': '84_Substance Abuse Disorder Inpatient Services_Benefit20_INN2_Coinsurance',
    'Benefit20_OON_Coinsurance': '83_Substance Abuse Disorder Inpatient Services_Benefit20_OON_Coinsurance',
    'Benefit21_INN_Copay': '126_Generic Drugs_Benefit21_INN_Copay',
    'Benefit21_INN2_Copay': '126_Generic Drugs_Benefit21_INN2_Copay',
    'Benefit21_OON_Copay': '126_Generic Drugs_Benefit21_OON_Copay',
    'Benefit21_INN_Coinsurance': '126_Generic Drugs_Benefit21_INN_Coinsurance',
    'Benefit21_INN2_Coinsurance': '126_Generic Drugs_Benefit21_INN2_Coinsurance',
    'Benefit21_OON_Coinsurance': '126_Generic Drugs_Benefit21_OON_Coinsurance',
    'Benefit22_INN_Copay': '127_Preferred Brand Drugs_Benefit22_INN_Copay',
    'Benefit22_INN2_Copay': '127_Preferred Brand Drugs_Benefit22_INN2_Copay',
    'Benefit22_OON_Copay': '127_Preferred Brand Drugs_Benefit22_OON_Copay',
    'Benefit22_INN_Coinsurance': '127_Preferred Brand Drugs_Benefit22_INN_Coinsurance',
    'Benefit22_INN2_Coinsurance': '127_Preferred Brand Drugs_Benefit22_INN2_Coinsurance',
    'Benefit22_OON_Coinsurance': '127_Preferred Brand Drugs_Benefit22_OON_Coinsurance',
    'Benefit23_INN_Copay': '128_Non-Preferred Brand Drugs_Benefit23_INN_Copay',
    'Benefit23_INN2_Copay': '128_Non-Preferred Brand Drugs_Benefit23_INN2_Copay',
    'Benefit23_OON_Copay': '128_Non-Preferred Brand Drugs_Benefit23_OON_Copay',
    'Benefit23_INN_Coinsurance': '128_Non-Preferred Brand Drugs_Benefit23_INN_Coinsurance',
    'Benefit23_INN2_Coinsurance': '128_Non-Preferred Brand Drugs_Benefit23_INN2_Coinsurance',
    'Benefit23_OON_Coinsurance': '128_Non-Preferred Brand Drugs_Benefit23_OON_Coinsurance',
    'Benefit24_INN_Copay': '129_Specialty Drugs_Benefit24_INN_Copay',
    'Benefit24_INN2_Copay': '129_Specialty Drugs_Benefit24_INN2_Copay',
    'Benefit24_OON_Copay': '129_Specialty Drugs_Benefit24_OON_Copay',
    'Benefit24_INN_Coinsurance': '129_Specialty Drugs_Benefit24_INN_Coinsurance',
    'Benefit24_INN2_Coinsurance': '129_Specialty Drugs_Benefit24_INN2_Coinsurance',
    'Benefit24_OON_Coinsurance': '129_Specialty Drugs_Benefit24_OON_Coinsurance',
    'Benefit25_INN_Copay': '176_Outpatient Rehabilitation Services_Benefit25_INN_Copay',
    'Benefit25_INN2_Copay': '176_Outpatient Rehabilitation Services_Benefit25_INN2_Copay',
    'Benefit25_OON_Copay': '176_Outpatient Rehabilitation Services_Benefit25_OON_Copay',
    'Benefit25_INN_Coinsurance': '176_Outpatient Rehabilitation Services_Benefit25_INN_Coinsurance',
    'Benefit25_INN2_Coinsurance': '176_Outpatient Rehabilitation Services_Benefit25_INN2_Coinsurance',
    'Benefit25_OON_Coinsurance': '176_Outpatient Rehabilitation Services_Benefit25_OON_Coinsurance',
    'Benefit26_INN_Copay': '164_Habilitation Services_Benefit26_INN_Copay',
    'Benefit26_INN2_Copay': '164_Habilitation Services_Benefit26_INN2_Copay',
    'Benefit26_OON_Copay': '164_Habilitation Services_Benefit26_OON_Copay',
    'Benefit26_INN_Coinsurance': '164_Habilitation Services_Benefit26_INN_Coinsurance',
    'Benefit26_INN2_Coinsurance': '164_Habilitation Services_Benefit26_INN2_Coinsurance',
    'Benefit26_OON_Coinsurance': '164_Habilitation Services_Benefit26_OON_Coinsurance',
    'Benefit27_INN_Copay': '29_Chiropractic Care_Benefit27_INN_Copay',
    'Benefit27_INN2_Copay': '29_Chiropractic Care_Benefit27_INN2_Copay',
    'Benefit27_OON_Copay': '29_Chiropractic Care_Benefit27_OON_Copay',
    'Benefit27_INN_Coinsurance': '29_Chiropractic Care_Benefit27_INN_Coinsurance',
    'Benefit27_INN2_Coinsurance': '29_Chiropractic Care_Benefit27_INN2_Coinsurance',
    'Benefit27_OON_Coinsurance': '29_Chiropractic Care_Benefit27_OON_Coinsurance',
    'Benefit28_INN_Copay': '173_Durable Medical Equipment_Benefit28_INN_Copay',
    'Benefit28_INN2_Copay': '173_Durable Medical Equipment_Benefit28_INN2_Copay',
    'Benefit28_OON_Copay': '173_Durable Medical Equipment_Benefit28_OON_Copay',
    'Benefit28_INN_Coinsurance': '173_Durable Medical Equipment_Benefit28_INN_Coinsurance',
    'Benefit28_INN2_Coinsurance': '173_Durable Medical Equipment_Benefit28_INN2_Coinsurance',
    'Benefit28_OON_Coinsurance': '173_Durable Medical Equipment_Benefit28_OON_Coinsurance',
    'Benefit29_INN_Copay': '173_Hearing Aids_Benefit29_INN_Copay',
    'Benefit29_INN2_Copay': '173_Hearing Aids_Benefit29_INN2_Copay',
    'Benefit29_OON_Copay': '173_Hearing Aids_Benefit29_OON_Copay',
    'Benefit29_INN_Coinsurance': '173_Hearing Aids_Benefit29_INN_Coinsurance',
    'Benefit29_INN2_Coinsurance': '173_Hearing Aids_Benefit29_INN2_Coinsurance',
    'Benefit29_OON_Coinsurance': '173_Hearing Aids_Benefit29_OON_Coinsurance',
    'Benefit30_INN_Copay': '63_Imaging (CT/PET Scans, MRIs)_Benefit30_INN_Copay',
    'Benefit30_INN2_Copay': '64_Imaging (CT/PET Scans, MRIs)_Benefit30_INN2_Copay',
    'Benefit30_OON_Copay': '63_Imaging (CT/PET Scans, MRIs)_Benefit30_OON_Copay',
    'Benefit30_INN_Coinsurance': '63_Imaging (CT/PET Scans, MRIs)_Benefit30_INN_Coinsurance',
    'Benefit30_INN2_Coinsurance': '64_Imaging (CT/PET Scans, MRIs)_Benefit30_INN2_Coinsurance',
    'Benefit30_OON_Coinsurance': '63_Imaging (CT/PET Scans, MRIs)_Benefit30_OON_Coinsurance',
    'Benefit31_INN_Copay': '1_Preventive Care/Screening/Immunization_Benefit31_INN_Copay',
    'Benefit31_INN2_Copay': '1_Preventive Care/Screening/Immunization_Benefit31_INN2_Copay',
    'Benefit31_OON_Copay': '1_Preventive Care/Screening/Immunization_Benefit31_OON_Copay',
    'Benefit31_INN_Coinsurance': '1_Preventive Care/Screening/Immunization_Benefit31_INN_Coinsurance',
    'Benefit31_INN2_Coinsurance': '1_Preventive Care/Screening/Immunization_Benefit31_INN2_Coinsurance',
    'Benefit31_OON_Coinsurance': '1_Preventive Care/Screening/Immunization_Benefit31_OON_Coinsurance',
    'Benefit32_INN_Copay': '12_Routine Foot Care_Benefit32_INN_Copay',
    'Benefit32_INN2_Copay': '12_Routine Foot Care_Benefit32_INN2_Copay',
    'Benefit32_OON_Copay': '12_Routine Foot Care_Benefit32_OON_Copay',
    'Benefit32_INN_Coinsurance': '12_Routine Foot Care_Benefit32_INN_Coinsurance',
    'Benefit32_INN2_Coinsurance': '12_Routine Foot Care_Benefit32_INN2_Coinsurance',
    'Benefit32_OON_Coinsurance': '12_Routine Foot Care_Benefit32_OON_Coinsurance',
    'Benefit33_INN_Copay': '186_Routine Eye Exam for Children_Benefit33_INN_Copay',
    'Benefit33_INN2_Copay': '186_Routine Eye Exam for Children_Benefit33_INN2_Copay',
    'Benefit33_OON_Copay': '186_Routine Eye Exam for Children_Benefit33_OON_Copay',
    'Benefit33_INN_Coinsurance': '186_Routine Eye Exam for Children_Benefit33_INN_Coinsurance',
    'Benefit33_INN2_Coinsurance': '186_Routine Eye Exam for Children_Benefit33_INN2_Coinsurance',
    'Benefit33_OON_Coinsurance': '186_Routine Eye Exam for Children_Benefit33_OON_Coinsurance',
    'Benefit34_INN_Copay': '187_Eye Glasses for Children_Benefit34_INN_Copay',
    'Benefit34_INN2_Copay': '187_Eye Glasses for Children_Benefit34_INN2_Copay',
    'Benefit34_OON_Copay': '187_Eye Glasses for Children_Benefit34_OON_Copay',
    'Benefit34_INN_Coinsurance': '187_Eye Glasses for Children_Benefit34_INN_Coinsurance',
    'Benefit34_INN2_Coinsurance': '187_Eye Glasses for Children_Benefit34_INN2_Coinsurance',
    'Benefit34_OON_Coinsurance': '187_Eye Glasses for Children_Benefit34_OON_Coinsurance',
    'Benefit35_INN_Copay': '209_Dental Check-Up for Children_Benefit35_INN_Copay',
    'Benefit35_INN2_Copay': '209_Dental Check-Up for Children_Benefit35_INN2_Copay',
    'Benefit35_OON_Copay': '209_Dental Check-Up for Children_Benefit35_OON_Copay',
    'Benefit35_INN_Coinsurance': '209_Dental Check-Up for Children_Benefit35_INN_Coinsurance',
    'Benefit35_INN2_Coinsurance': '209_Dental Check-Up for Children_Benefit35_INN2_Coinsurance',
    'Benefit35_OON_Coinsurance': '209_Dental Check-Up for Children_Benefit35_OON_Coinsurance',
    'Benefit36_INN_Copay': '176_Rehabilitative Speech Therapy_Benefit36_INN_Copay',
    'Benefit36_INN2_Copay': '176_Rehabilitative Speech Therapy_Benefit36_INN2_Copay',
    'Benefit36_OON_Copay': '176_Rehabilitative Speech Therapy_Benefit36_OON_Copay',
    'Benefit36_INN_Coinsurance': '176_Rehabilitative Speech Therapy_Benefit36_INN_Coinsurance',
    'Benefit36_INN2_Coinsurance': '176_Rehabilitative Speech Therapy_Benefit36_INN2_Coinsurance',
    'Benefit36_OON_Coinsurance': '176_Rehabilitative Speech Therapy_Benefit36_OON_Coinsurance',
    'Benefit37_INN_Copay': '175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_INN_Copay',
    'Benefit37_INN2_Copay': '175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_INN2_Copay',
    'Benefit37_OON_Copay': '175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_OON_Copay',
    'Benefit37_INN_Coinsurance': '175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_INN_Coinsurance',
    'Benefit37_INN2_Coinsurance': '175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_INN2_Coinsurance',
    'Benefit37_OON_Coinsurance': '175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_OON_Coinsurance',
    'Benefit38_INN_Copay': '1_Well Baby Visits and Care_Benefit38_INN_Copay',
    'Benefit38_INN2_Copay': '1_Well Baby Visits and Care_Benefit38_INN2_Copay',
    'Benefit38_OON_Copay': '1_Well Baby Visits and Care_Benefit38_OON_Copay',
    'Benefit38_INN_Coinsurance': '1_Well Baby Visits and Care_Benefit38_INN_Coinsurance',
    'Benefit38_INN2_Coinsurance': '1_Well Baby Visits and Care_Benefit38_INN2_Coinsurance',
    'Benefit38_OON_Coinsurance': '1_Well Baby Visits and Care_Benefit38_OON_Coinsurance',
    'Benefit39_INN_Copay': '18_Laboratory Outpatient and Professional Services_Benefit39_INN_Copay',
    'Benefit39_INN2_Copay': '18_Laboratory Outpatient and Professional Services_Benefit39_INN2_Copay',
    'Benefit39_OON_Copay': '18_Laboratory Outpatient and Professional Services_Benefit39_OON_Copay',
    'Benefit39_INN_Coinsurance': '18_Laboratory Outpatient and Professional Services_Benefit39_INN_Coinsurance',
    'Benefit39_INN2_Coinsurance': '18_Laboratory Outpatient and Professional Services_Benefit39_INN2_Coinsurance',
    'Benefit39_OON_Coinsurance': '18_Laboratory Outpatient and Professional Services_Benefit39_OON_Coinsurance',
    'Benefit40_INN_Copay': '19_X-rays and Diagnostic Imaging_Benefit40_INN_Copay',
    'Benefit40_INN2_Copay': '19_X-rays and Diagnostic Imaging_Benefit40_INN2_Copay',
    'Benefit40_OON_Copay': '19_X-rays and Diagnostic Imaging_Benefit40_OON_Copay',
    'Benefit40_INN_Coinsurance': '19_X-rays and Diagnostic Imaging_Benefit40_INN_Coinsurance',
    'Benefit40_INN2_Coinsurance': '19_X-rays and Diagnostic Imaging_Benefit40_INN2_Coinsurance',
    'Benefit40_OON_Coinsurance': '19_X-rays and Diagnostic Imaging_Benefit40_OON_Coinsurance',
    'Benefit41_INN_Copay': '210_Basic Dental Care – Child_Benefit41_INN_Copay',
    'Benefit41_INN2_Copay': '210_Basic Dental Care – Child_Benefit41_INN2_Copay',
    'Benefit41_OON_Copay': '210_Basic Dental Care – Child_Benefit41_OON_Copay',
    'Benefit41_INN_Coinsurance': '210_Basic Dental Care – Child_Benefit41_INN_Coinsurance',
    'Benefit41_INN2_Coinsurance': '210_Basic Dental Care – Child_Benefit41_INN2_Coinsurance',
    'Benefit41_OON_Coinsurance': '210_Basic Dental Care – Child_Benefit41_OON_Coinsurance',
    'Benefit42_INN_Copay': '216_Orthodontia – Child_Benefit42_INN_Copay',
    'Benefit42_INN2_Copay': '216_Orthodontia – Child_Benefit42_INN2_Copay',
    'Benefit42_OON_Copay': '216_Orthodontia – Child_Benefit42_OON_Copay',
    'Benefit42_INN_Coinsurance': '216_Orthodontia – Child_Benefit42_INN_Coinsurance',
    'Benefit42_INN2_Coinsurance': '216_Orthodontia – Child_Benefit42_INN2_Coinsurance',
    'Benefit42_OON_Coinsurance': '216_Orthodontia – Child_Benefit42_OON_Coinsurance',
    'Benefit43_INN_Copay': '214_Major Dental Care – Child_Benefit43_INN_Copay',
    'Benefit43_INN2_Copay': '214_Major Dental Care – Child_Benefit43_INN2_Copay',
    'Benefit43_OON_Copay': '214_Major Dental Care – Child_Benefit43_OON_Copay',
    'Benefit43_INN_Coinsurance': '214_Major Dental Care – Child_Benefit43_INN_Coinsurance',
    'Benefit43_INN2_Coinsurance': '214_Major Dental Care – Child_Benefit43_INN2_Coinsurance',
    'Benefit43_OON_Coinsurance': '214_Major Dental Care – Child_Benefit43_OON_Coinsurance',
    'Benefit44_INN_Copay': '87_Transplant_Benefit44_INN_Copay',
    'Benefit44_INN2_Copay': '88_Transplant_Benefit44_INN2_Copay',
    'Benefit44_OON_Copay': '87_Transplant_Benefit44_OON_Copay',
    'Benefit44_INN_Coinsurance': '87_Transplant_Benefit44_INN_Coinsurance',
    'Benefit44_INN2_Coinsurance': '88_Transplant_Benefit44_INN2_Coinsurance',
    'Benefit44_OON_Coinsurance': '87_Transplant_Benefit44_OON_Coinsurance',
    'Benefit45_INN_Copay': '12_Accidental Dental_Benefit45_INN_Copay',
    'Benefit45_INN2_Copay': '12_Accidental Dental_Benefit45_INN2_Copay',
    'Benefit45_OON_Copay': '12_Accidental Dental_Benefit45_OON_Copay',
    'Benefit45_INN_Coinsurance': '12_Accidental Dental_Benefit45_INN_Coinsurance',
    'Benefit45_INN2_Coinsurance': '12_Accidental Dental_Benefit45_INN2_Coinsurance',
    'Benefit45_OON_Coinsurance': '12_Accidental Dental_Benefit45_OON_Coinsurance',
    'Benefit46_INN_Copay': '73_Dialysis_Benefit46_INN_Copay',
    'Benefit46_INN2_Copay': '73_Dialysis_Benefit46_INN2_Copay',
    'Benefit46_OON_Copay': '73_Dialysis_Benefit46_OON_Copay',
    'Benefit46_INN_Coinsurance': '73_Dialysis_Benefit46_INN_Coinsurance',
    'Benefit46_INN2_Coinsurance': '73_Dialysis_Benefit46_INN2_Coinsurance',
    'Benefit46_OON_Coinsurance': '73_Dialysis_Benefit46_OON_Coinsurance',
    'Benefit47_INN_Copay': '15_Allergy Testing_Benefit47_INN_Copay',
    'Benefit47_INN2_Copay': '15_Allergy Testing_Benefit47_INN2_Copay',
    'Benefit47_OON_Copay': '15_Allergy Testing_Benefit47_OON_Copay',
    'Benefit47_INN_Coinsurance': '15_Allergy Testing_Benefit47_INN_Coinsurance',
    'Benefit47_INN2_Coinsurance': '15_Allergy Testing_Benefit47_INN2_Coinsurance',
    'Benefit47_OON_Coinsurance': '15_Allergy Testing_Benefit47_OON_Coinsurance',
    'Benefit48_INN_Copay': '69_Chemotherapy_Benefit48_INN_Copay',
    'Benefit48_INN2_Copay': '69_Chemotherapy_Benefit48_INN2_Copay',
    'Benefit48_OON_Copay': '69_Chemotherapy_Benefit48_OON_Copay',
    'Benefit48_INN_Coinsurance': '69_Chemotherapy_Benefit48_INN_Coinsurance',
    'Benefit48_INN2_Coinsurance': '69_Chemotherapy_Benefit48_INN2_Coinsurance',
    'Benefit48_OON_Coinsurance': '69_Chemotherapy_Benefit48_OON_Coinsurance',
    'Benefit49_INN_Copay': '69_Radiation_Benefit49_INN_Copay',
    'Benefit49_INN2_Copay': '69_Radiation_Benefit49_INN2_Copay',
    'Benefit49_OON_Copay': '69_Radiation_Benefit49_OON_Copay',
    'Benefit49_INN_Coinsurance': '69_Radiation_Benefit49_INN_Coinsurance',
    'Benefit49_INN2_Coinsurance': '69_Radiation_Benefit49_INN2_Coinsurance',
    'Benefit49_OON_Coinsurance': '69_Radiation_Benefit49_OON_Coinsurance',
    'Benefit50_INN_Copay': '12_Diabetes Education_Benefit50_INN_Copay',
    'Benefit50_INN2_Copay': '12_Diabetes Education_Benefit50_INN2_Copay',
    'Benefit50_OON_Copay': '12_Diabetes Education_Benefit50_OON_Copay',
    'Benefit50_INN_Coinsurance': '12_Diabetes Education_Benefit50_INN_Coinsurance',
    'Benefit50_INN2_Coinsurance': '12_Diabetes Education_Benefit50_INN2_Coinsurance',
    'Benefit50_OON_Coinsurance': '12_Diabetes Education_Benefit50_OON_Coinsurance',
    'Benefit51_INN_Copay': '183_Prosthetic Devices_Benefit51_INN_Copay',
    'Benefit51_INN2_Copay': '183_Prosthetic Devices_Benefit51_INN2_Copay',
    'Benefit51_OON_Copay': '183_Prosthetic Devices_Benefit51_OON_Copay',
    'Benefit51_INN_Coinsurance': '183_Prosthetic Devices_Benefit51_INN_Coinsurance',
    'Benefit51_INN2_Coinsurance': '183_Prosthetic Devices_Benefit51_INN2_Coinsurance',
    'Benefit51_OON_Coinsurance': '183_Prosthetic Devices_Benefit51_OON_Coinsurance',
    'Benefit52_INN_Copay': '69_Infusion Therapy_Benefit52_INN_Copay',
    'Benefit52_INN2_Copay': '69_Infusion Therapy_Benefit52_INN2_Copay',
    'Benefit52_OON_Copay': '69_Infusion Therapy_Benefit52_OON_Copay',
    'Benefit52_INN_Coinsurance': '69_Infusion Therapy_Benefit52_INN_Coinsurance',
    'Benefit52_INN2_Coinsurance': '69_Infusion Therapy_Benefit52_INN2_Coinsurance',
    'Benefit52_OON_Coinsurance': '69_Infusion Therapy_Benefit52_OON_Coinsurance',
    'Benefit53_INN_Copay': '47_Treatment for Temporomandibular Joint Disorders_Benefit53_INN_Copay',
    'Benefit53_INN2_Copay': '48_Treatment for Temporomandibular Joint Disorders_Benefit53_INN2_Copay',
    'Benefit53_OON_Copay': '47_Treatment for Temporomandibular Joint Disorders_Benefit53_OON_Copay',
    'Benefit53_INN_Coinsurance': '47_Treatment for Temporomandibular Joint Disorders_Benefit53_INN_Coinsurance',
    'Benefit53_INN2_Coinsurance': '48_Treatment for Temporomandibular Joint Disorders_Benefit53_INN2_Coinsurance',
    'Benefit53_OON_Coinsurance': '47_Treatment for Temporomandibular Joint Disorders_Benefit53_OON_Coinsurance',
    'Benefit54_INN_Copay': '11_Nutritional Counseling_Benefit54_INN_Copay',
    'Benefit54_INN2_Copay': '11_Nutritional Counseling_Benefit54_INN2_Copay',
    'Benefit54_OON_Copay': '11_Nutritional Counseling_Benefit54_OON_Copay',
    'Benefit54_INN_Coinsurance': '11_Nutritional Counseling_Benefit54_INN_Coinsurance',
    'Benefit54_INN2_Coinsurance': '11_Nutritional Counseling_Benefit54_INN2_Coinsurance',
    'Benefit54_OON_Coinsurance': '11_Nutritional Counseling_Benefit54_OON_Coinsurance',
    'Benefit55_INN_Copay': '79_Reconstructive Surgery_Benefit55_INN_Copay',
    'Benefit55_INN2_Copay': '80_Reconstructive Surgery_Benefit55_INN2_Copay',
    'Benefit55_OON_Copay': '79_Reconstructive Surgery_Benefit55_OON_Copay',
    'Benefit55_INN_Coinsurance': '79_Reconstructive Surgery_Benefit55_INN_Coinsurance',
    'Benefit55_INN2_Coinsurance': '80_Reconstructive Surgery_Benefit55_INN2_Coinsurance',
    'Benefit55_OON_Coinsurance': '79_Reconstructive Surgery_Benefit55_OON_Coinsurance',
    'Benefit56_INN_Copay': '79_Clinical Trials_Benefit56_INN_Copay',
    'Benefit56_INN2_Copay': '80_Clinical Trials_Benefit56_INN2_Copay',
    'Benefit56_OON_Copay': '79_Clinical Trials_Benefit56_OON_Copay',
    'Benefit56_INN_Coinsurance': '79_Clinical Trials_Benefit56_INN_Coinsurance',
    'Benefit56_INN2_Coinsurance': '80_Clinical Trials_Benefit56_INN2_Coinsurance',
    'Benefit56_OON_Coinsurance': '79_Clinical Trials_Benefit56_OON_Coinsurance',
    'Benefit57_INN_Copay': '128_Inherited Metabolic Disorder - PKU_Benefit57_INN_Copay',
    'Benefit57_INN2_Copay': '128_Inherited Metabolic Disorder - PKU_Benefit57_INN2_Copay',
    'Benefit57_OON_Copay': '128_Inherited Metabolic Disorder - PKU_Benefit57_OON_Copay',
    'Benefit57_INN_Coinsurance': '128_Inherited Metabolic Disorder - PKU_Benefit57_INN_Coinsurance',
    'Benefit57_INN2_Coinsurance': '128_Inherited Metabolic Disorder - PKU_Benefit57_INN2_Coinsurance',
    'Benefit57_OON_Coinsurance': '128_Inherited Metabolic Disorder - PKU_Benefit57_OON_Coinsurance',
    'Benefit58_INN_Copay': '134_Dental Anesthesia_Benefit58_INN_Copay',
    'Benefit58_INN2_Copay': '134_Dental Anesthesia_Benefit58_INN2_Copay',
    'Benefit58_OON_Copay': '134_Dental Anesthesia_Benefit58_OON_Copay',
    'Benefit58_INN_Coinsurance': '134_Dental Anesthesia_Benefit58_INN_Coinsurance',
    'Benefit58_INN2_Coinsurance': '134_Dental Anesthesia_Benefit58_INN2_Coinsurance',
    'Benefit58_OON_Coinsurance': '134_Dental Anesthesia_Benefit58_OON_Coinsurance',
    'Benefit59_INN_Copay': '49_Bone Marrow Testing_Benefit59_INN_Copay',
    'Benefit59_INN2_Copay': '49_Bone Marrow Testing_Benefit59_INN2_Copay',
    'Benefit59_OON_Copay': '49_Bone Marrow Testing_Benefit59_OON_Copay',
    'Benefit59_INN_Coinsurance': '49_Bone Marrow Testing_Benefit59_INN_Coinsurance',
    'Benefit59_INN2_Coinsurance': '49_Bone Marrow Testing_Benefit59_INN2_Coinsurance',
    'Benefit59_OON_Coinsurance': '49_Bone Marrow Testing_Benefit59_OON_Coinsurance',
    'Benefit60_INN_Copay': '79_Newborn Services Other_Benefit60_INN_Copay',
    'Benefit60_INN2_Copay': '80_Newborn Services Other_Benefit60_INN2_Copay',
    'Benefit60_OON_Copay': '79_Newborn Services Other_Benefit60_OON_Copay',
    'Benefit60_INN_Coinsurance': '79_Newborn Services Other_Benefit60_INN_Coinsurance',
    'Benefit60_INN2_Coinsurance': '80_Newborn Services Other_Benefit60_INN2_Coinsurance',
    'Benefit60_OON_Coinsurance': '79_Newborn Services Other_Benefit60_OON_Coinsurance',
    'Benefit61_INN_Copay': '156_Applied Behavior Analysis Based Therapies_Benefit61_INN_Copay',
    'Benefit61_INN2_Copay': '156_Applied Behavior Analysis Based Therapies_Benefit61_INN2_Copay',
    'Benefit61_OON_Copay': '156_Applied Behavior Analysis Based Therapies_Benefit61_OON_Copay',
    'Benefit61_INN_Coinsurance': '156_Applied Behavior Analysis Based Therapies_Benefit61_INN_Coinsurance',
    'Benefit61_INN2_Coinsurance': '156_Applied Behavior Analysis Based Therapies_Benefit61_INN2_Coinsurance',
    'Benefit61_OON_Coinsurance': '156_Applied Behavior Analysis Based Therapies_Benefit61_OON_Coinsurance',

}
qhp_costshare_df = qhp_costshare_df.rename(columns=column_mapping)
qhp_costshare_df = qhp_costshare_df.drop(0)
# Template sheet name for copying the structure
template_sheet_name = 'data 1'

merged_df = pd.merge(
    hp_a2ch_df,
    qhp_costshare_df,
    left_on='HIOS Plan ID* (Standard Component + Variant)',
    right_on='0HIOS_Plan_ID_With_Variant',
    how='outer'
)
# If an index column is created, drop it
if 'Unnamed: 0' in merged_df.columns:
    merged_df.drop(columns=['Unnamed: 0'], inplace=True)

columns_to_drop = ['0HIOS_Plan_ID_With_Variant']
merged_df = merged_df.drop(columns=columns_to_drop)

merged_df['hios_id'] = merged_df['HIOS Plan ID* (Standard Component + Variant)'].str.split('-').str[0].str.strip()

# Check if the template sheet exists in the workbook
if template_sheet_name not in workbook.sheetnames:
    raise ValueError(f"Template sheet '{template_sheet_name}' not found in the workbook.")


# Remove completely empty rows from mindsection_df
mindsection_df.dropna(how='all', inplace=True)

# Helper function to set cell values while maintaining formatting
def set_cell_value(sheet, cell_reference, value):
    font = Font(name='Arial', size=11)
    for merged_range in sheet.merged_cells.ranges:
        if cell_reference in merged_range:
            min_row, min_col, _, _ = merged_range.bounds
            top_left_cell = sheet.cell(row=min_row, column=min_col)
            top_left_cell.value = value
            top_left_cell.font = font
            return
    cell = sheet[cell_reference]
    cell.value = value
    cell.font = font

# Function to process Benefit Package sheets
def process_benefit_package_sheet(value, sheet_name):
    # Check if the sheet exists in the workbook
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        # Create a new sheet with the same structure as the template
        template_sheet = workbook[template_sheet_name]
        sheet = workbook.copy_worksheet(template_sheet)
        sheet.title = sheet_name

    # Extract the required data
    BenefitPackage = value['Benefit Package']
    BenefitInformation = value['Benefit Information']

    # Check if the BenefitPackage and BenefitInformation are not empty
    if BenefitPackage.empty or BenefitInformation.empty:
        return False

    # Extracting required information
    first_row_value = BenefitPackage.iloc[0]
    hios_plan_id = first_row_value['HIOS Plan ID*(Standard Component)']
    ke = hios_plan_id[:5]
    state = hios_plan_id[5:7]

    market_coverage = mindsection_df['Market Coverage*'].iloc[0] if 'Market Coverage*' in mindsection_df.columns else 'Individual'
    dental_only_plan = mindsection_df['Dental Only Plan*'].iloc[0] if 'Dental Only Plan*' in mindsection_df.columns else 'No'

    # Set values for the specified fields
    set_cell_value(sheet, 'B2', ke)
    set_cell_value(sheet, 'B3', state)
    set_cell_value(sheet, 'B4', market_coverage)
    set_cell_value(sheet, 'B5', dental_only_plan)

    # Perform the left join with an indicator column
    left_join = pd.merge(
        mindsection_df,
        BenefitPackage.rename(columns={'HIOS Plan ID*(Standard Component)': 'HIOS Plan ID* (Standard Component)'}),
        on='HIOS Plan ID* (Standard Component)',
        how='left',
        indicator=True
    )

    # Filter the rows where the match happened (both columns match)
    filtered_left_join = left_join[left_join['_merge'] == 'both'].drop(columns='_merge')

    # Drop additional columns as needed
    columns_to_drop = ['lastModificationDate', 'Refresh Date', 'Plan Marketing Name*_x', 'Level of Coverage*_x', 'Plan Type*_x']
    filtered_left_join = filtered_left_join.drop(columns=columns_to_drop)

    # Rename columns as needed
    columns_to_rename = {'Plan Marketing Name*_y': 'Plan Marketing Name*', 'Level of Coverage*_y': 'Level of Coverage*', 'Plan Type*_y': 'Plan Type*'}
    filtered_left_join = filtered_left_join.rename(columns=columns_to_rename)

    required_columns = [
        "HIOS Plan ID* (Standard Component)",
        "Plan Marketing Name*", "HIOS Product ID*",
        "Network ID*", "Service Area ID*", "Formulary ID*", "New/Existing Plan?*",
        "Plan Type*", "Level of Coverage*", "Design Type*", "Unique Plan Design?*",
        "QHP/Non QHP*", "Notice Required for Pregnancy*", "Plan Level Exclusions",
        "Limited Cost Sharing Plan Variation - Est Advanced Payment",
        "Does this plan offer Composite Rating?*", "Child-Only Offering*",
        "Child Only Plan ID", "Tobacco Wellness Program Offered*",
        "Disease Management Programs Offered", "EHB Percent of Total Premium*",
        "EHB Apportionment for Pediatric Dental", "Guaranteed vs_ Estimated Rate",
        "Plan Effective Date*", "Plan Expiration Date", "Out of Country Coverage*",
        "Out of Country Coverage Description", "Out of Service Area Coverage*",
        "Out of Service Area Coverage Description", "National Network*"
    ]
    # Create a new DataFrame with only the specified column
    existing_columns = [col for col in required_columns if col in filtered_left_join.columns]
    filtered_left_join = filtered_left_join[existing_columns]
    # # Ensure there are no duplicate rows
    # filtered_left_join = filtered_left_join.drop_duplicates()

    # Write DataFrame to the sheet starting from row 8, column 1 without changing formatting
    start_row = 8
    current_row = start_row
    font = Font(name='Arial', size=11)
    for row_idx, row in filtered_left_join.iterrows():
        for col_idx, value in enumerate(row):
            dest_cell = sheet.cell(row=current_row, column=col_idx + 1)
            dest_cell.value = value
            dest_cell.font = font
            if col_idx :  # Assuming 'Plan Marketing Name*' is the second column and 'Disease Management Programs Offered' is the third column
                dest_cell.alignment = Alignment(wrap_text=True, vertical='top')
                sheet.column_dimensions[dest_cell.column_letter].width =30   # Adjust width as needed
            else:
                dest_cell.alignment = Alignment(vertical='top')
        current_row += 1

    # Remove the first column from BenefitInformation
    BenefitInformation = BenefitInformation.drop(BenefitInformation.columns[0], axis=1)

    # Populate Benefit Information data starting from C59 without changing formatting
    benefit_start_row = 60
    benefit_start_col = 3
    benefit_data = BenefitInformation

    for row_idx, row in benefit_data.iterrows():
        for col_idx, value in enumerate(row):
            dest_cell = sheet.cell(row=benefit_start_row + row_idx, column=benefit_start_col + col_idx)
            dest_cell.value = value
            dest_cell.font = font
            if col_idx :  # Assuming 'Plan Marketing Name*' is the second column and 'Disease Management Programs Offered' is the third column
                dest_cell.alignment = Alignment(wrap_text=True, vertical='top')
                sheet.column_dimensions[dest_cell.column_letter].width =20   # Adjust width as needed
            else:
                dest_cell.alignment = Alignment(vertical='top')
    return True

# Function to create and populate a Cost Share Variances sheet
def create_and_populate_cost_share_variances_sheet(base_sheet_name, new_sheet_name, data):
    global data_written  # To modify the outer scope variable

    if new_sheet_name in workbook.sheetnames:
        sheet = workbook[new_sheet_name]
    else:
        if base_sheet_name in workbook.sheetnames:
            base_sheet = workbook[base_sheet_name]
            sheet = workbook.copy_worksheet(base_sheet)
            sheet.title = new_sheet_name
        else:
            raise ValueError(f"Base sheet '{base_sheet_name}' not found in the workbook.")


    start_row = 4
    font = Font(name='Arial', size=11)


# Uncomment and adjust the list of columns you want to select
    data = data[[
        'HIOS Plan ID* (Standard Component + Variant)', 'Plan Marketing Name*', 
        'Level of Coverage* (Metal Level)', 'CSR Variation Type*', 'Issuer Actuarial Value', 
        'AV Calculator Output Number*', 'Medical & Drug Deductibles Integrated?*', 
        'Medical & Drug Maximum Out of Pocket Integrated?*', 'Is a Referral Required for Specialist?*', 
        'Specialist(s) Requiring a Referral', 'Multiple In Network Tiers?*', '1st Tier Utilization*', 
        '2nd Tier Utilization', 'SBC_Maternity_Deductible', 'SBC_Maternity_Copay', 
        'SBC_Maternity_Coinsurance', 'SBC_Maternity_Limits', 'SBC_Diabetes_Deductible', 
        'SBC_Diabetes_Copay', 'SBC_Diabetes_Coinsurance', 'SBC_Diabetes_Limits', 
        'SBC_FootFracture_Deductible', 'SBC_FootFracture_Copay', 'SBC_FootFracture_Coinsurance', 
        'SBC_FootFracture_Limits', 'Maximum Out of Pocket for Medical EHB Benefits: In Network Individual', 
        'Maximum Out of Pocket for Medical EHB Benefits: In Network Family', 
        'Maximum Out of Pocket for Medical EHB Benefits: In Network (Tier 2) Individual', 
        'Maximum Out of Pocket for Medical EHB Benefits: In Network (Tier 2) Family', 
        'Maximum Out of Pocket for Medical EHB Benefits: Out of Network Individual', 
        'Maximum Out of Pocket for Medical EHB Benefits: Out of Network Family', 
        'Maximum Out of Pocket for Medical EHB Benefits: Combined In/Out Network Individual', 
        'Maximum Out of Pocket for Medical EHB Benefits: Combined In/Out Network Family', 
        'Maximum Out of Pocket for Drug EHB Benefits: In Network Individual', 
        'Maximum Out of Pocket for Drug EHB Benefits: In Network Family', 
        'Maximum Out of Pocket for Drug EHB Benefits: In Network (Tier 2) Individual', 
        'Maximum Out of Pocket for Drug EHB Benefits: In Network (Tier 2) Family', 
        'Maximum Out of Pocket for Drug EHB Benefits: Out of Network Individual', 
        'Maximum Out of Pocket for Drug EHB Benefits: Out of Network Family', 
        'Maximum Out of Pocket for Drug EHB Benefits: Combined In/Out Network Individual', 
        'Maximum Out of Pocket for Drug EHB Benefits: Combined In/Out Network Family', 
        'Maximum Out of Pocket for Medical and Drug EHB Benefits (Total): In Network Individual', 
        'Maximum Out of Pocket for Medical and Drug EHB Benefits (Total): In Network Family', 
        'Maximum Out of Pocket for Medical and Drug EHB Benefits (Total): In Network (Tier 2) Individual', 
        'Maximum Out of Pocket for Medical and Drug EHB Benefits (Total): In Network (Tier 2) Family', 
        'Maximum Out of Pocket for Medical and Drug EHB Benefits (Total): Out of Network Individual', 
        'Maximum Out of Pocket for Medical and Drug EHB Benefits (Total): Out of Network Family', 
        'Maximum Out of Pocket for Medical and Drug EHB Benefits (Total) Combined In/Out Network Individual', 
        'Maximum Out of Pocket for Medical and Drug EHB Benefits (Total): Combined In/Out Network Family', 
        'Medical EHB Deductible: In Network Individual', 'Medical EHB Deductible: In Network Family', 
        'Medical EHB Deductible: In Network Default Coinsurance', 'Medical EHB Deductible: In Network (Tier 2) Individual', 
        'Medical EHB Deductible: In Network (Tier 2) Family', 'Medical EHB Deductible: In Network (Tier 2) Default Coinsurance', 
        'Medical EHB Deductible: Out of Network Individual', 'Medical EHB Deductible: Out of Network Family', 
        'Medical EHB Deductible: Combined In/Out Network Individual', 'Medical EHB Deductible: Combined In/Out Network Family', 
        'Drug EHB Deductible: In Network Individual', 'Drug EHB Deductible: In Network Family', 
        'Drug EHB Deductible: In Network Default Coinsurance', 'Drug EHB Deductible: In Network (Tier 2) Individual', 
        'Drug EHB Deductible: In Network (Tier 2) Family', 'Drug EHB Deductible: In Network (Tier 2) Default Coinsurance', 
        'Drug EHB Deductible: Out of Network Individual', 'Drug EHB Deductible: Out of Network Family', 
        'Drug EHB Deductible: Combined In/Out Network Individual', 'Drug EHB Deductible: Combined In/Out Network Family', 
        'Combined Medical and Drug EHB Deductible: In Network Individual', 'Combined Medical and Drug EHB Deductible: In Network Family', 
        'Combined Medical and Drug EHB Deductible: In Network Default Coinsurance', 
        'Combined Medical and Drug EHB Deductible: In Network (Tier 2) Individual', 
        'Combined Medical and Drug EHB Deductible: In Network (Tier 2) Family', 
        'Combined Medical and Drug EHB Deductible: In Network (Tier 2) Default Coinsurance', 
        'Combined Medical and Drug EHB Deductible: Out of Network Individual', 
        'Combined Medical and Drug EHB Deductible: Out of Network Family', 
        'Combined Medical and Drug EHB Deductible: Combined In/Out Network Individual', 
        'Combined Medical and Drug EHB Deductible: Combined In/Out Network Family', 'HSA Eligible*', 
        'HSA/HRA Employer Contribution', 'HSA/HRA Employer Contribution Amount', 'Maximum Coinsurance for Specialty Drugs', 
        'Maximum Number of Days for Charging an Inpatient Copay?', 
        'Begin Primary Care Cost-Sharing After a Set Number of Visits?', 
        'Begin Primary Care Deductible/Coinsurance After a Set Number of Copays?',
          "2_Primary Care Visit to Treat an Injury or Illness_Benefit1_INN_Copay",
    "2_Primary Care Visit to Treat an Injury or Illness_Benefit1_INN2_Copay",
    "2_Primary Care Visit to Treat an Injury or Illness_Benefit1_OON_Copay",
    "2_Primary Care Visit to Treat an Injury or Illness_Benefit1_INN_Coinsurance",
    "2_Primary Care Visit to Treat an Injury or Illness_Benefit1_INN2_Coinsurance",
    "2_Primary Care Visit to Treat an Injury or Illness_Benefit1_OON_Coinsurance",
    "12_Specialist Visit_Benefit2_INN_Copay",
    "12_Specialist Visit_Benefit2_INN2_Copay",
    "12_Specialist Visit_Benefit2_OON_Copay",
    "12_Specialist Visit_Benefit2_INN_Coinsurance",
    "12_Specialist Visit_Benefit2_INN2_Coinsurance",
    "12_Specialist Visit_Benefit2_OON_Coinsurance",
    "2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_INN_Copay",
    "2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_INN2_Copay",
    "2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_OON_Copay",
    "2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_INN_Coinsurance",
    "2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_INN2_Coinsurance",
    "2_Other Practitioner Office Visit (Nurse, Physician Assistant)_Benefit3_OON_Coinsurance",
    "47_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_INN_Copay",
    "48_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_INN2_Copay",
    "47_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_OON_Copay",
    "47_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_INN_Coinsurance",
    "48_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_INN2_Coinsurance",
    "47_Outpatient Facility Fee (e.g., Ambulatory Surgery Center)_Benefit4_OON_Coinsurance",
    "50_Outpatient Surgery Physician/Surgical Services_Benefit5_INN_Copay",
    "50_Outpatient Surgery Physician/Surgical Services_Benefit5_INN2_Copay",
    "50_Outpatient Surgery Physician/Surgical Services_Benefit5_OON_Copay",
    "50_Outpatient Surgery Physician/Surgical Services_Benefit5_INN_Coinsurance",
    "50_Outpatient Surgery Physician/Surgical Services_Benefit5_INN2_Coinsurance",
    "50_Outpatient Surgery Physician/Surgical Services_Benefit5_OON_Coinsurance",
    "254_Hospice Services_Benefit6_INN_Copay",
    "254_Hospice Services_Benefit6_INN2_Copay",
    "254_Hospice Services_Benefit6_OON_Copay",
    "254_Hospice Services_Benefit6_INN_Coinsurance",
    "254_Hospice Services_Benefit6_INN2_Coinsurance",
    "254_Hospice Services_Benefit6_OON_Coinsurance",
    "253_Private-Duty Nursing_Benefit7_INN_Copay",
    "253_Private-Duty Nursing_Benefit7_INN2_Copay",
    "253_Private-Duty Nursing_Benefit7_OON_Copay",
    "253_Private-Duty Nursing_Benefit7_INN_Coinsurance",
    "253_Private-Duty Nursing_Benefit7_INN2_Coinsurance",
    "253_Private-Duty Nursing_Benefit7_OON_Coinsurance",
    "33_Urgent Care Centers or Facilities_Benefit8_INN_Copay",
    "33_Urgent Care Centers or Facilities_Benefit8_INN2_Copay",
    "33_Urgent Care Centers or Facilities_Benefit8_OON_Copay",
    "33_Urgent Care Centers or Facilities_Benefit8_INN_Coinsurance",
    "33_Urgent Care Centers or Facilities_Benefit8_INN2_Coinsurance",
    "33_Urgent Care Centers or Facilities_Benefit8_OON_Coinsurance",
    "98_Home Health Care Services_Benefit9_INN_Copay",
    "98_Home Health Care Services_Benefit9_INN2_Copay",
    "98_Home Health Care Services_Benefit9_OON_Copay",
    "98_Home Health Care Services_Benefit9_INN_Coinsurance",
    "98_Home Health Care Services_Benefit9_INN2_Coinsurance",
    "98_Home Health Care Services_Benefit9_OON_Coinsurance",
    "43_Emergency Room Services_Benefit10_INN_Copay",
    "43_Emergency Room Services_Benefit10_INN2_Copay",
    "43_Emergency Room Services_Benefit10_OON_Copay",
    "43_Emergency Room Services_Benefit10_INN_Coinsurance",
    "43_Emergency Room Services_Benefit10_INN2_Coinsurance",
    "43_Emergency Room Services_Benefit10_OON_Coinsurance",
    "103_Emergency Transportation/Ambulance_Benefit11_INN_Copay",
    "103_Emergency Transportation/Ambulance_Benefit11_INN2_Copay",
    "103_Emergency Transportation/Ambulance_Benefit11_OON_Copay",
    "103_Emergency Transportation/Ambulance_Benefit11_INN_Coinsurance",
    "103_Emergency Transportation/Ambulance_Benefit11_INN2_Coinsurance",
    "103_Emergency Transportation/Ambulance_Benefit11_OON_Coinsurance",
    "79_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_INN_Copay",
    "80_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_INN2_Copay",
    "79_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_OON_Copay",
    "79_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_INN_Coinsurance",
    "80_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_INN2_Coinsurance",
    "79_Inpatient Hospital Services (e.g., Hospital Stay)_Benefit12_OON_Coinsurance",
    "93_Inpatient Physician and Surgical Services_Benefit13_INN_Copay",
    "93_Inpatient Physician and Surgical Services_Benefit13_INN2_Copay",
    "93_Inpatient Physician and Surgical Services_Benefit13_OON_Copay",
    "93_Inpatient Physician and Surgical Services_Benefit13_INN_Coinsurance",
    "93_Inpatient Physician and Surgical Services_Benefit13_INN2_Coinsurance",
    "93_Inpatient Physician and Surgical Services_Benefit13_OON_Coinsurance",
    "81_Skilled Nursing Facility_Benefit14_INN_Copay",
    "82_Skilled Nursing Facility_Benefit14_INN2_Copay",
    "81_Skilled Nursing Facility_Benefit14_OON_Copay",
    "81_Skilled Nursing Facility_Benefit14_INN_Coinsurance",
    "82_Skilled Nursing Facility_Benefit14_INN2_Coinsurance",
    "81_Skilled Nursing Facility_Benefit14_OON_Coinsurance",
    "6_Prenatal and Postnatal Care_Benefit15_INN_Copay",
    "6_Prenatal and Postnatal Care_Benefit15_INN2_Copay",
    "6_Prenatal and Postnatal Care_Benefit15_OON_Copay",
    "6_Prenatal and Postnatal Care_Benefit15_INN_Coinsurance",
    "6_Prenatal and Postnatal Care_Benefit15_INN2_Coinsurance",
    "6_Prenatal and Postnatal Care_Benefit15_OON_Coinsurance",
    "79_Delivery and All Inpatient Services for Maternity Care_Benefit16_INN_Copay",
    "80_Delivery and All Inpatient Services for Maternity Care_Benefit16_INN2_Copay",
    "79_Delivery and All Inpatient Services for Maternity Care_Benefit16_OON_Copay",
    "79_Delivery and All Inpatient Services for Maternity Care_Benefit16_INN_Coinsurance",
    "80_Delivery and All Inpatient Services for Maternity Care_Benefit16_INN2_Coinsurance",
    "79_Delivery and All Inpatient Services for Maternity Care_Benefit16_OON_Coinsurance",
    "246_Mental/Behavioral Health Outpatient Services_Benefit17_INN_Copay",
    "246_Mental/Behavioral Health Outpatient Services_Benefit17_INN2_Copay",
    "246_Mental/Behavioral Health Outpatient Services_Benefit17_OON_Copay",
    "246_Mental/Behavioral Health Outpatient Services_Benefit17_INN_Coinsurance",
    "246_Mental/Behavioral Health Outpatient Services_Benefit17_INN2_Coinsurance",
    "246_Mental/Behavioral Health Outpatient Services_Benefit17_OON_Coinsurance",
    "83_Mental/Behavioral Health Inpatient Services_Benefit18_INN_Copay",
    "84_Mental/Behavioral Health Inpatient Services_Benefit18_INN2_Copay",
    "83_Mental/Behavioral Health Inpatient Services_Benefit18_OON_Copay",
    # "83_Mental/Behavioral Health Inpatient Services_Benefit18_INN_Coinsurance",
    # "84_Mental/Behavioral Health Inpatient Services_Benefit18_INN2_Coinsurance",
    # "83_Mental/Behavioral Health Inpatient Services_Benefit18_OON_Coinsurance",
    # "246_Substance Abuse Disorder Outpatient Services_Benefit19_INN_Copay",
    # "246_Substance Abuse Disorder Outpatient Services_Benefit19_INN2_Copay",
    # "246_Substance Abuse Disorder Outpatient Services_Benefit19_OON_Copay",
    # "246_Substance Abuse Disorder Outpatient Services_Benefit19_INN_Coinsurance",
    # "246_Substance Abuse Disorder Outpatient Services_Benefit19_INN2_Coinsurance",
    # "246_Substance Abuse Disorder Outpatient Services_Benefit19_OON_Coinsurance",
    # "83_Substance Abuse Disorder Inpatient Services_Benefit20_INN_Copay",
    # "84_Substance Abuse Disorder Inpatient Services_Benefit20_INN2_Copay",
    # "83_Substance Abuse Disorder Inpatient Services_Benefit20_OON_Copay",
    # "83_Substance Abuse Disorder Inpatient Services_Benefit20_INN_Coinsurance",
    # "84_Substance Abuse Disorder Inpatient Services_Benefit20_INN2_Coinsurance",
    # "83_Substance Abuse Disorder Inpatient Services_Benefit20_OON_Coinsurance",
    # "126_Generic Drugs_Benefit21_INN_Copay",
    # "126_Generic Drugs_Benefit21_INN2_Copay",
    # "126_Generic Drugs_Benefit21_OON_Copay",
    # "126_Generic Drugs_Benefit21_INN_Coinsurance",
    # "126_Generic Drugs_Benefit21_INN2_Coinsurance",
    # "126_Generic Drugs_Benefit21_OON_Coinsurance",
    # "127_Preferred Brand Drugs_Benefit22_INN_Copay",
    # "127_Preferred Brand Drugs_Benefit22_INN2_Copay",
    # "127_Preferred Brand Drugs_Benefit22_OON_Copay",
    # "127_Preferred Brand Drugs_Benefit22_INN_Coinsurance",
    # "127_Preferred Brand Drugs_Benefit22_INN2_Coinsurance",
    # "127_Preferred Brand Drugs_Benefit22_OON_Coinsurance",
    # "128_Non-Preferred Brand Drugs_Benefit23_INN_Copay",
    # "128_Non-Preferred Brand Drugs_Benefit23_INN2_Copay",
    # "128_Non-Preferred Brand Drugs_Benefit23_OON_Copay",
    # "128_Non-Preferred Brand Drugs_Benefit23_INN_Coinsurance",
    # "128_Non-Preferred Brand Drugs_Benefit23_INN2_Coinsurance",
    # "128_Non-Preferred Brand Drugs_Benefit23_OON_Coinsurance",
    # "129_Specialty Drugs_Benefit24_INN_Copay",
    # "129_Specialty Drugs_Benefit24_INN2_Copay",
    # "129_Specialty Drugs_Benefit24_OON_Copay",
    # "129_Specialty Drugs_Benefit24_INN_Coinsurance",
    # "129_Specialty Drugs_Benefit24_INN2_Coinsurance",
    # "129_Specialty Drugs_Benefit24_OON_Coinsurance",
    # "176_Outpatient Rehabilitation Services_Benefit25_INN_Copay",
    # "176_Outpatient Rehabilitation Services_Benefit25_INN2_Copay",
    # "176_Outpatient Rehabilitation Services_Benefit25_OON_Copay",
    # "176_Outpatient Rehabilitation Services_Benefit25_INN_Coinsurance",
    # "176_Outpatient Rehabilitation Services_Benefit25_INN2_Coinsurance",
    # "176_Outpatient Rehabilitation Services_Benefit25_OON_Coinsurance",
    # "164_Habilitation Services_Benefit26_INN_Copay",
    # "164_Habilitation Services_Benefit26_INN2_Copay",
    # "164_Habilitation Services_Benefit26_OON_Copay",
    # "164_Habilitation Services_Benefit26_INN_Coinsurance",
    # "164_Habilitation Services_Benefit26_INN2_Coinsurance",
    # "164_Habilitation Services_Benefit26_OON_Coinsurance",
    # "29_Chiropractic Care_Benefit27_INN_Copay",
    # "29_Chiropractic Care_Benefit27_INN2_Copay",
    # "29_Chiropractic Care_Benefit27_OON_Copay",
    # "29_Chiropractic Care_Benefit27_INN_Coinsurance",
    # "29_Chiropractic Care_Benefit27_INN2_Coinsurance",
    # "29_Chiropractic Care_Benefit27_OON_Coinsurance",
    # "173_Durable Medical Equipment_Benefit28_INN_Copay",
    # "173_Durable Medical Equipment_Benefit28_INN2_Copay",
    # "173_Durable Medical Equipment_Benefit28_OON_Copay",
    # "173_Durable Medical Equipment_Benefit28_INN_Coinsurance",
    # "173_Durable Medical Equipment_Benefit28_INN2_Coinsurance",
    # "173_Durable Medical Equipment_Benefit28_OON_Coinsurance",
    # "173_Hearing Aids_Benefit29_INN_Copay",
    # "173_Hearing Aids_Benefit29_INN2_Copay",
    # "173_Hearing Aids_Benefit29_OON_Copay",
    # "173_Hearing Aids_Benefit29_INN_Coinsurance",
    # "173_Hearing Aids_Benefit29_INN2_Coinsurance",
    # "173_Hearing Aids_Benefit29_OON_Coinsurance",
    # "63_Imaging (CT/PET Scans, MRIs)_Benefit30_INN_Copay",
    # "64_Imaging (CT/PET Scans, MRIs)_Benefit30_INN2_Copay",
    # "63_Imaging (CT/PET Scans, MRIs)_Benefit30_OON_Copay",
    # "63_Imaging (CT/PET Scans, MRIs)_Benefit30_INN_Coinsurance",
    # "64_Imaging (CT/PET Scans, MRIs)_Benefit30_INN2_Coinsurance",
    # "63_Imaging (CT/PET Scans, MRIs)_Benefit30_OON_Coinsurance",
    # "1_Preventive Care/Screening/Immunization_Benefit31_INN_Copay",
    # "1_Preventive Care/Screening/Immunization_Benefit31_INN2_Copay",
    # "1_Preventive Care/Screening/Immunization_Benefit31_OON_Copay",
    # "1_Preventive Care/Screening/Immunization_Benefit31_INN_Coinsurance",
    # "1_Preventive Care/Screening/Immunization_Benefit31_INN2_Coinsurance",
    # "1_Preventive Care/Screening/Immunization_Benefit31_OON_Coinsurance",
    # "12_Routine Foot Care_Benefit32_INN_Copay",
    # "12_Routine Foot Care_Benefit32_INN2_Copay",
    # "12_Routine Foot Care_Benefit32_OON_Copay",
    # "12_Routine Foot Care_Benefit32_INN_Coinsurance",
    # "12_Routine Foot Care_Benefit32_INN2_Coinsurance",
    # "12_Routine Foot Care_Benefit32_OON_Coinsurance",
    # "186_Routine Eye Exam for Children_Benefit33_INN_Copay",
    # "186_Routine Eye Exam for Children_Benefit33_INN2_Copay",
    # "186_Routine Eye Exam for Children_Benefit33_OON_Copay",
    # "186_Routine Eye Exam for Children_Benefit33_INN_Coinsurance",
    # "186_Routine Eye Exam for Children_Benefit33_INN2_Coinsurance",
    # "186_Routine Eye Exam for Children_Benefit33_OON_Coinsurance",
    # "187_Eye Glasses for Children_Benefit34_INN_Copay",
    # "187_Eye Glasses for Children_Benefit34_INN2_Copay",
    # "187_Eye Glasses for Children_Benefit34_OON_Copay",
    # "187_Eye Glasses for Children_Benefit34_INN_Coinsurance",
    # "187_Eye Glasses for Children_Benefit34_INN2_Coinsurance",
    # "187_Eye Glasses for Children_Benefit34_OON_Coinsurance",
    # "209_Dental Check-Up for Children_Benefit35_INN_Copay",
    # "209_Dental Check-Up for Children_Benefit35_INN2_Copay",
    # "209_Dental Check-Up for Children_Benefit35_OON_Copay",
    # "209_Dental Check-Up for Children_Benefit35_INN_Coinsurance",
    # "209_Dental Check-Up for Children_Benefit35_INN2_Coinsurance",
    # "209_Dental Check-Up for Children_Benefit35_OON_Coinsurance",
    # "176_Rehabilitative Speech Therapy_Benefit36_INN_Copay",
    # "176_Rehabilitative Speech Therapy_Benefit36_INN2_Copay",
    # "176_Rehabilitative Speech Therapy_Benefit36_OON_Copay",
    # "176_Rehabilitative Speech Therapy_Benefit36_INN_Coinsurance",
    # "176_Rehabilitative Speech Therapy_Benefit36_INN2_Coinsurance",
    # "176_Rehabilitative Speech Therapy_Benefit36_OON_Coinsurance",
    # "175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_INN_Copay",
    # "175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_INN2_Copay",
    # "175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_OON_Copay",
    # "175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_INN_Coinsurance",
    # "175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_INN2_Coinsurance",
    # "175_Rehabilitative Occupational and Rehabilitative Physical Therapy_Benefit37_OON_Coinsurance",
    # "1_Well Baby Visits and Care_Benefit38_INN_Copay",
    # "1_Well Baby Visits and Care_Benefit38_INN2_Copay",
    # "1_Well Baby Visits and Care_Benefit38_OON_Copay",
    # "1_Well Baby Visits and Care_Benefit38_INN_Coinsurance",
    # "1_Well Baby Visits and Care_Benefit38_INN2_Coinsurance",
    # "1_Well Baby Visits and Care_Benefit38_OON_Coinsurance",
    # "18_Laboratory Outpatient and Professional Services_Benefit39_INN_Copay",
    # "18_Laboratory Outpatient and Professional Services_Benefit39_INN2_Copay",
    # "18_Laboratory Outpatient and Professional Services_Benefit39_OON_Copay",
    # "18_Laboratory Outpatient and Professional Services_Benefit39_INN_Coinsurance",
    # "18_Laboratory Outpatient and Professional Services_Benefit39_INN2_Coinsurance",
    # "18_Laboratory Outpatient and Professional Services_Benefit39_OON_Coinsurance",
    # "19_X-rays and Diagnostic Imaging_Benefit40_INN_Copay",
    # "19_X-rays and Diagnostic Imaging_Benefit40_INN2_Copay",
    # "19_X-rays and Diagnostic Imaging_Benefit40_OON_Copay",
    # "19_X-rays and Diagnostic Imaging_Benefit40_INN_Coinsurance",
    # "19_X-rays and Diagnostic Imaging_Benefit40_INN2_Coinsurance",
    # "19_X-rays and Diagnostic Imaging_Benefit40_OON_Coinsurance",
    # "210_Basic Dental Care – Child_Benefit41_INN_Copay",
    # "210_Basic Dental Care – Child_Benefit41_INN2_Copay",
    # "210_Basic Dental Care – Child_Benefit41_OON_Copay",
    # "210_Basic Dental Care – Child_Benefit41_INN_Coinsurance",
    # "210_Basic Dental Care – Child_Benefit41_INN2_Coinsurance",
    # "210_Basic Dental Care – Child_Benefit41_OON_Coinsurance",
    # "216_Orthodontia – Child_Benefit42_INN_Copay",
    # "216_Orthodontia – Child_Benefit42_INN2_Copay",
    # "216_Orthodontia – Child_Benefit42_OON_Copay",
    # "216_Orthodontia – Child_Benefit42_INN_Coinsurance",
    # "216_Orthodontia – Child_Benefit42_INN2_Coinsurance",
    # "216_Orthodontia – Child_Benefit42_OON_Coinsurance",
    # "214_Major Dental Care – Child_Benefit43_INN_Copay",
    # "214_Major Dental Care – Child_Benefit43_INN2_Copay",
    # "214_Major Dental Care – Child_Benefit43_OON_Copay",
    # "214_Major Dental Care – Child_Benefit43_INN_Coinsurance",
    # "214_Major Dental Care – Child_Benefit43_INN2_Coinsurance",
    # "214_Major Dental Care – Child_Benefit43_OON_Coinsurance",
    # "87_Transplant_Benefit44_INN_Copay",
    # "88_Transplant_Benefit44_INN2_Copay",
    # "87_Transplant_Benefit44_OON_Copay",
    # "87_Transplant_Benefit44_INN_Coinsurance",
    # "88_Transplant_Benefit44_INN2_Coinsurance",
    # "87_Transplant_Benefit44_OON_Coinsurance",
    # "12_Accidental Dental_Benefit45_INN_Copay",
    # "12_Accidental Dental_Benefit45_INN2_Copay",
    # "12_Accidental Dental_Benefit45_OON_Copay",
    # "12_Accidental Dental_Benefit45_INN_Coinsurance",
    # "12_Accidental Dental_Benefit45_INN2_Coinsurance",
    # "12_Accidental Dental_Benefit45_OON_Coinsurance",
    # "73_Dialysis_Benefit46_INN_Copay",
    # "73_Dialysis_Benefit46_INN2_Copay",
    # "73_Dialysis_Benefit46_OON_Copay",
    # "73_Dialysis_Benefit46_INN_Coinsurance",
    # "73_Dialysis_Benefit46_INN2_Coinsurance",
    # "73_Dialysis_Benefit46_OON_Coinsurance",
    # "15_Allergy Testing_Benefit47_INN_Copay",
    # "15_Allergy Testing_Benefit47_INN2_Copay",
    # "15_Allergy Testing_Benefit47_OON_Copay",
    # "15_Allergy Testing_Benefit47_INN_Coinsurance",
    # "15_Allergy Testing_Benefit47_INN2_Coinsurance",
    # "15_Allergy Testing_Benefit47_OON_Coinsurance",
    # "69_Chemotherapy_Benefit48_INN_Copay",
    # "69_Chemotherapy_Benefit48_INN2_Copay",
    # "69_Chemotherapy_Benefit48_OON_Copay",
    # "69_Chemotherapy_Benefit48_INN_Coinsurance",
    # "69_Chemotherapy_Benefit48_INN2_Coinsurance",
    # "69_Chemotherapy_Benefit48_OON_Coinsurance",
    # "69_Radiation_Benefit49_INN_Copay",
    # "69_Radiation_Benefit49_INN2_Copay",
    # "69_Radiation_Benefit49_OON_Copay",
    # "69_Radiation_Benefit49_INN_Coinsurance",
    # "69_Radiation_Benefit49_INN2_Coinsurance",
    # "69_Radiation_Benefit49_OON_Coinsurance",
    # "12_Diabetes Education_Benefit50_INN_Copay",
    # "12_Diabetes Education_Benefit50_INN2_Copay",
    # "12_Diabetes Education_Benefit50_OON_Copay",
    # "12_Diabetes Education_Benefit50_INN_Coinsurance",
    # "12_Diabetes Education_Benefit50_INN2_Coinsurance",
    # "12_Diabetes Education_Benefit50_OON_Coinsurance",
    # "183_Prosthetic Devices_Benefit51_INN_Copay",
    # "183_Prosthetic Devices_Benefit51_INN2_Copay",
    # "183_Prosthetic Devices_Benefit51_OON_Copay",
    # "183_Prosthetic Devices_Benefit51_INN_Coinsurance",
    # "183_Prosthetic Devices_Benefit51_INN2_Coinsurance",
    # "183_Prosthetic Devices_Benefit51_OON_Coinsurance",
    # "69_Infusion Therapy_Benefit52_INN_Copay",
    # "69_Infusion Therapy_Benefit52_INN2_Copay",
    # "69_Infusion Therapy_Benefit52_OON_Copay",
    # "69_Infusion Therapy_Benefit52_INN_Coinsurance",
    # "69_Infusion Therapy_Benefit52_INN2_Coinsurance",
    # "69_Infusion Therapy_Benefit52_OON_Coinsurance",
    # "47_Treatment for Temporomandibular Joint Disorders_Benefit53_INN_Copay",
    # "48_Treatment for Temporomandibular Joint Disorders_Benefit53_INN2_Copay",
    # "47_Treatment for Temporomandibular Joint Disorders_Benefit53_OON_Copay",
    # "47_Treatment for Temporomandibular Joint Disorders_Benefit53_INN_Coinsurance",
    # "48_Treatment for Temporomandibular Joint Disorders_Benefit53_INN2_Coinsurance",
    # "47_Treatment for Temporomandibular Joint Disorders_Benefit53_OON_Coinsurance",
    # "11_Nutritional Counseling_Benefit54_INN_Copay",
    # "11_Nutritional Counseling_Benefit54_INN2_Copay",
    # "11_Nutritional Counseling_Benefit54_OON_Copay",
    # "11_Nutritional Counseling_Benefit54_INN_Coinsurance",
    # "11_Nutritional Counseling_Benefit54_INN2_Coinsurance",
    # "11_Nutritional Counseling_Benefit54_OON_Coinsurance",
    # "79_Reconstructive Surgery_Benefit55_INN_Copay",
    # "80_Reconstructive Surgery_Benefit55_INN2_Copay",
    # "79_Reconstructive Surgery_Benefit55_OON_Copay",
    # "79_Reconstructive Surgery_Benefit55_INN_Coinsurance",
    # "80_Reconstructive Surgery_Benefit55_INN2_Coinsurance",
    # "79_Reconstructive Surgery_Benefit55_OON_Coinsurance",
    # "79_Clinical Trials_Benefit56_INN_Copay",
    # "80_Clinical Trials_Benefit56_INN2_Copay",
    # "79_Clinical Trials_Benefit56_OON_Copay",
    # "79_Clinical Trials_Benefit56_INN_Coinsurance",
    # "80_Clinical Trials_Benefit56_INN2_Coinsurance",
    # "79_Clinical Trials_Benefit56_OON_Coinsurance",
    # "128_Inherited Metabolic Disorder - PKU_Benefit57_INN_Copay",
    # "128_Inherited Metabolic Disorder - PKU_Benefit57_INN2_Copay",
    # "128_Inherited Metabolic Disorder - PKU_Benefit57_OON_Copay",
    # "128_Inherited Metabolic Disorder - PKU_Benefit57_INN_Coinsurance",
    # "128_Inherited Metabolic Disorder - PKU_Benefit57_INN2_Coinsurance",
    # "128_Inherited Metabolic Disorder - PKU_Benefit57_OON_Coinsurance",
    # "134_Dental Anesthesia_Benefit58_INN_Copay",
    # "134_Dental Anesthesia_Benefit58_INN2_Copay",
    # "134_Dental Anesthesia_Benefit58_OON_Copay",
    # "134_Dental Anesthesia_Benefit58_INN_Coinsurance",
    # "134_Dental Anesthesia_Benefit58_INN2_Coinsurance",
    # "134_Dental Anesthesia_Benefit58_OON_Coinsurance",
    # "49_Bone Marrow Testing_Benefit59_INN_Copay",
    # "49_Bone Marrow Testing_Benefit59_INN2_Copay",
    # "49_Bone Marrow Testing_Benefit59_OON_Copay",
    # "49_Bone Marrow Testing_Benefit59_INN_Coinsurance",
    # "49_Bone Marrow Testing_Benefit59_INN2_Coinsurance",
    # "49_Bone Marrow Testing_Benefit59_OON_Coinsurance",
    # "79_Newborn Services Other_Benefit60_INN_Copay",
    # "80_Newborn Services Other_Benefit60_INN2_Copay",
    # "79_Newborn Services Other_Benefit60_OON_Copay",
    # "79_Newborn Services Other_Benefit60_INN_Coinsurance",
    # "80_Newborn Services Other_Benefit60_INN2_Coinsurance",
    # "79_Newborn Services Other_Benefit60_OON_Coinsurance",
    # "156_Applied Behavior Analysis Based Therapies_Benefit61_INN_Copay",
    # "156_Applied Behavior Analysis Based Therapies_Benefit61_INN2_Copay",
    # "156_Applied Behavior Analysis Based Therapies_Benefit61_OON_Copay",
    # "156_Applied Behavior Analysis Based Therapies_Benefit61_INN_Coinsurance",
    # "156_Applied Behavior Analysis Based Therapies_Benefit61_INN2_Coinsurance",
    # "156_Applied Behavior Analysis Based Therapies_Benefit61_OON_Coinsurance"


    ]]


    for row_idx, row in data.iterrows():
        for col_idx, value in enumerate(row):
            dest_cell = sheet.cell(row=start_row + row_idx, column=col_idx + 1)
            dest_cell.value = value
            dest_cell.font = font
            if col_idx :  # Assuming 'Plan Marketing Name*' is the second column and 'Disease Management Programs Offered' is the third column
                dest_cell.alignment = Alignment(wrap_text=True, vertical='top')
                sheet.column_dimensions[dest_cell.column_letter].width =40   # Adjust width as needed
            else:
                dest_cell.alignment = Alignment(vertical='top')
    sheet.freeze_panes = 'F2'

    # Remove empty rows
    max_row = sheet.max_row
    for row in range(max_row, start_row - 1, -1):
        if all(cell.value is None for cell in sheet[row]):
            sheet.delete_rows(row)

    data_written = True  # Set flag to True when writing data

    return True

for key, value in dfs.items():
    if process_benefit_package_sheet(value, key):
        data_written = True
        hios_plan_ids = value['Benefit Package']['HIOS Plan ID*(Standard Component)'].unique()
        filtered_cost_share_data = merged_df[merged_df['hios_id'].isin(hios_plan_ids)].copy()
        filtered_cost_share_data=filtered_cost_share_data.drop([9,8,7,6,5,4,3,2,1,0])
        create_and_populate_cost_share_variances_sheet("data 2", f"Cost Share Variances {key.split()[-1]}", filtered_cost_share_data)

if data_written:
    # Remove 'ram' and 'ram1' sheets if they exist
    for sheet_name in ['ram', 'data 1','data 2']:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    # Save the workbook with a new name
    new_file_path = 'output//output_PY2025PlansBenefitsTemplate.xlsm'
    workbook.save(new_file_path)
    print(f"Workbook saved as '{new_file_path}'")
else:
    print("No data was written.")
