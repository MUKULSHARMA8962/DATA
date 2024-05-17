import pandas as pd
import openpyxl
from openpyxl.styles import Font
from lnd_file import dfs
# Load the workbook with VBA macros enabled
file_path = 'PY2025PlansBenefitsTemplate.xlsm'
workbook = openpyxl.load_workbook(file_path, keep_vba=True)

# Load the CSV file into a DataFrame
filename = "qhp_midsection (3) (1).csv"
mindsection_df = pd.read_csv(filename)

hp_a2chfilename = "qhp_a2ch (3) (1).csv"
hp_a2ch_df = pd.read_csv(hp_a2chfilename)


qhp_costsharefilename = "qhp_costshare (2) (1).csv"
qhp_costshare_df = pd.read_csv(qhp_costsharefilename)
# Template sheet name for copying the structure
template_sheet_name = 'Benefits Package 1'


merged_df = pd.merge(
    hp_a2ch_df, 
    qhp_costshare_df, 
    left_on='HIOS Plan ID* (Standard Component + Variant)', 
    right_on='hios_plan_id_with_variant', 
    how='outer',
    suffixes=('_hp_a2ch', '_qhp_costshare')
)
merged_df['hios_id'] = merged_df['HIOS Plan ID* (Standard Component + Variant)'].str.split('-').str[0].str.strip()




# Check if the template sheet exists in the workbook
if template_sheet_name not in workbook.sheetnames:
    raise ValueError(f"Template sheet '{template_sheet_name}' not found in the workbook.")

# Map keys to the expected sheet names
key_mapping = {f'Benefit PKG {i}': f'Benefits Package {i}' for i in range(1, len(dfs) + 1)}

# Print the key mapping for debugging
print(f"Key Mapping: {key_mapping}")

# Flag to track if any data was written
data_written = False

# Remove completely empty rows
mindsection_df.dropna(how='all', inplace=True)


# mindsection_df['HIOS Plan ID* (Standard Component)']

# id aye use ek varibale store karna 
# then  merged_df   ek ccolume hai hios_id   

# data match hota hai use   use data ko upolad  Cost Share Variances   all sheet mai 


# Helper function to set cell values while maintaining formatting
def set_cell_value(sheet, cell_reference, value):
    font = Font(name='Arial', size=11)
    for merged_range in sheet.merged_cells.ranges:
        if cell_reference in merged_range:
            min_row, min_col, _, _ = merged_range.bounds
            top_left_cell = sheet.cell(row=min_row, column=min_col)
            top_left_cell.value = value
            top_left_cell.font = font
            return
    cell = sheet[cell_reference]
    cell.value = value
    cell.font = font

# Function to process Benefit Package sheets
def process_benefit_package_sheet(key, value, sheet_name):
    print(f"Processing key: {key}, mapped to sheet: {sheet_name}")
    
    # Check if the sheet exists in the workbook
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        # Create a new sheet with the same structure as the template
        template_sheet = workbook[template_sheet_name]
        sheet = workbook.copy_worksheet(template_sheet)
        sheet.title = sheet_name
    
    # Extract the required data
    BenefitPackage = value['Benefit Package']
    BenefitInformation = value['Benefit Information']

    # Extracting required information
    first_row_value = BenefitPackage.loc[0]
    hios_plan_id = first_row_value['HIOS Plan ID*(Standard Component)']
    ke = hios_plan_id[:5]
    state = hios_plan_id[5:7]

    market_coverage = mindsection_df['Market Coverage*'].iloc[0] if 'Market Coverage*' in mindsection_df.columns else 'Individual'
    dental_only_plan = mindsection_df['Dental Only Plan*'].iloc[0] if 'Dental Only Plan*' in mindsection_df.columns else 'No'

    # Set values for the specified fields
    set_cell_value(sheet, 'B2', ke)
    set_cell_value(sheet, 'B3', state)
    set_cell_value(sheet, 'B4', market_coverage)
    set_cell_value(sheet, 'B5', dental_only_plan)

    # Perform the left join with an indicator column
    left_join = pd.merge(
        mindsection_df,
        BenefitPackage.rename(columns={'HIOS Plan ID*(Standard Component)': 'HIOS Plan ID* (Standard Component)'}),
        on='HIOS Plan ID* (Standard Component)',
        how='left',
        indicator=True
    )
    
    # Filter the rows where the match happened (both columns match)
    filtered_left_join = left_join[left_join['_merge'] == 'both'].drop(columns='_merge')
    
    # Drop additional columns as needed
    columns_to_drop = ['lastModificationDate', 'Refresh Date', 'Plan Marketing Name*_x', 'Level of Coverage*_x', 'Plan Type*_x']
    filtered_left_join = filtered_left_join.drop(columns=columns_to_drop)

    # Rename columns as needed
    columns_to_rename = {'Plan Marketing Name*_y': 'Plan Marketing Name*', 'Level of Coverage*_y': 'Level of Coverage*', 'Plan Type*_y': 'Plan Type*'}
    filtered_left_join = filtered_left_join.rename(columns=columns_to_rename)

    # Ensure there are no duplicate rows
    filtered_left_join = filtered_left_join.drop_duplicates()

    # Write DataFrame to the sheet starting from row 8, column 1 without changing formatting
    start_row = 8
    current_row = start_row
    font = Font(name='Arial', size=11)
    for row_idx, row in filtered_left_join.iterrows():
        if not row.isnull().all():  # Check if the row is not empty
            for col_idx, value in enumerate(row):
                dest_cell = sheet.cell(row=current_row, column=col_idx + 1)
                dest_cell.value = value
                dest_cell.font = font
            current_row += 1

    # Remove the first column from BenefitInformation
    BenefitInformation = BenefitInformation.drop(BenefitInformation.columns[0], axis=1)

    # Populate Benefit Information data starting from C59 without changing formatting
    benefit_start_row = 60
    benefit_start_col = 3
    benefit_data = BenefitInformation.head(20)
    
    for row_idx, row in benefit_data.iterrows():
        for col_idx, value in enumerate(row):
            dest_cell = sheet.cell(row=benefit_start_row + row_idx, column=benefit_start_col + col_idx)
            dest_cell.value = value
            dest_cell.font = font
    
    return True

# Function to create Cost Share Variances sheets if not already present
def create_cost_share_variances_sheet(base_sheet_name, new_sheet_name):
    if new_sheet_name in workbook.sheetnames:
        print(f"Sheet {new_sheet_name} already exists, skipping creation.")
        return workbook[new_sheet_name]
    
    if base_sheet_name in workbook.sheetnames:
        base_sheet = workbook[base_sheet_name]
        new_sheet = workbook.copy_worksheet(base_sheet)
        new_sheet.title = new_sheet_name
        print(f"Created new sheet: {new_sheet_name} based on {base_sheet_name}")
    else:
        raise ValueError(f"Base sheet '{base_sheet_name}' not found in the workbook.")
    
    return new_sheet

# Iterate over each key, value pair in dfs
for key, value in dfs.items():
    # Map the key to the expected sheet name
    if key in key_mapping:
        sheet_name = key_mapping[key]
        data_written = process_benefit_package_sheet(key, value, sheet_name)
        
        # Create corresponding Cost Share Variances sheet
        cost_share_variances_sheet_name = f"Cost Share Variances {sheet_name.split()[-1]}"
        create_cost_share_variances_sheet("Cost Share Variances 1", cost_share_variances_sheet_name)

if data_written:
    # Save the workbook with a new name
    new_file_path = 'Modified_PY2025PlansBenefitsTemplate.xlsm'
    workbook.save(new_file_path)
    print(f"Workbook saved as '{new_file_path}'")
else:
    print("No data was written.")
