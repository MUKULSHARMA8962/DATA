import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from lnd_file import dfs

# Create a new workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove the default sheet created with the workbook

# Define the borders and fill colors
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

header_fill = PatternFill(start_color="D9EDBF", end_color="D9EDBF", fill_type="solid")  # Green color for headers
instruction_fill = PatternFill(start_color="CCCC99", end_color="CCCC99", fill_type="solid")  # Light tan color for instructions

# File name
filename = "qhp_midsection (3) (1).csv"

# Reading the CSV file into a DataFrame
mindsection_df = pd.read_csv(filename)

# Iterate over each key, value pair in dfs
for key, value in dfs.items():
    # Extracting the required data
    BenefitPackage = value['Benefit Package']
    BenefitPackage['HIOS Plan ID*(Standard Component)']
    # Select only the 'HIOS Plan ID*(Standard Component)' column from BenefitPackage
    BenefitPackage_subset = BenefitPackage[['HIOS Plan ID*(Standard Component)']]
    
    # Perform the left join with an indicator column
    left_join = pd.merge(
        mindsection_df,
        BenefitPackage.rename(columns={'HIOS Plan ID*(Standard Component)': 'HIOS Plan ID* (Standard Component)'}),
        on='HIOS Plan ID* (Standard Component)',
        how='left',
        indicator=True
    )
    
    # Filter the rows where the match happened (both columns match)
    filtered_left_join = left_join[left_join['_merge'] == 'both'].drop(columns='_merge')
    
    # Drop additional columns as needed
    columns_to_drop = ['lastModificationDate', 'Refresh Date','Plan Marketing Name*_x',"Level of Coverage*_x","Plan Type*_x"]
    filtered_left_join = filtered_left_join.drop(columns=columns_to_drop)

    # Rename columns as needed
    columns_to_rename = {'Plan Marketing Name*_y': 'Plan Marketing Name*', 'Level of Coverage*_y': 'Level of Coverage*','Plan Type*_y':'Plan Type*'}
    filtered_left_join = filtered_left_join.rename(columns=columns_to_rename)

    BenefitInformation = value['Benefit Information']
    first_row_value = BenefitPackage.loc[0]
    hios_plan_id = first_row_value['HIOS Plan ID*(Standard Component)']
    
    if pd.notna(hios_plan_id):
        ke = hios_plan_id[:5]  # Extract HIOS Plan ID
        state = hios_plan_id[5:7]  # Extract State

        # Create a new sheet for each BenefitPackage
        sheet_name = f"{key}"
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name}_{len(wb.sheetnames)}"
        ws = wb.create_sheet(title=sheet_name)

        # Set the title of the sheet and freeze panes
        ws.freeze_panes = 'A8'  # Freeze the rows above A8

        # Set initial instructional content with colors
        instructions = [
            ("2025 Plans & Benefits Template v14.0", "A1:C1"),
            ("HIOS Issuer ID*: 32753", "A2:A2"),
            ("Issuer State*: MO", "A3:A3"),
            ("Market Coverage*: Individual", "A4:A4"),
            ("Dental Only Plan*: No", "A5:A5"),
            ("To use this template, please review the user guide and instructions. All fields with an asterisk (*) are required", "D1:K1"),
            ("you will need to save the latest version of the add-in file (PlansBenefitsAddInPY25.xlam) on your machine.", "D2:K2"),
            ("To create the cost share variance worksheet and enter the cost sharing amounts for both individual and SHOP (small group) markets, use the Create Cost Share Variances macro.", "D3:K3"),
            ("To create additional Benefits Package worksheets, use the Create New Benefits Package macro.", "D4:K4"),
            ("To populate the benefits on the Benefits Package worksheet with your State EHB Standards, use the Refresh EHB macro.", "D5:K5"),
            ("Plane Identifiers ","A6:F6"),
            ("Plan Attributes","G6:U6"),
            ("Stand Alone Dental Only ","V6:W6"),
            ("Plan Dates	","X6:Y6"),
            ("Geographi Coverage","Z6:AD6"),

        ]

     
    # Ranges that need specific formatting
    specific_ranges = [
        "A6:F6", "G6:U6", "V6:W6", "X6:Y6", "Z6:AD6"
    ]   

    for text, cell_range in instructions:
        ws.merge_cells(cell_range)
        
        start_cell = cell_range.split(':')[0]
        cell = ws[start_cell]
        cell.value = text
        
        # Apply alignment and bold font with increased size to the title cell
        if cell_range == "A1:C1":  # Check if it's the title cell range
            for row in ws[cell_range]:
                for c in row:
                    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    c.font = Font(bold=True, size=14)  # Set font to bold and size 14
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # Apply fill color and specific formatting to the specified ranges
        if cell_range in specific_ranges:
            for row in ws[cell_range]:
                for c in row:
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    c.font = Font(bold=True, size=15)  # Set font to bold and size 12
                    c.fill = header_fill  # Apply header fill color
                    c.border = thin_border


        # Headers and data in Column A and B
        headers_data = [
            ("HIOS Issuer ID*", f"{ke}"),
            ("Issuer State*", f"{state}"),
            ("Market Coverage*", "Individual"),
            ("Dental Only Plan*", "No")
        ]

        for idx, (header, value) in enumerate(headers_data, start=2):
            header_cell = ws[f"A{idx}"]
            header_cell.value = header
            header_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border

            value_cell = ws[f"B{idx}"]
            value_cell.value = value
            value_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            value_cell.border = thin_border

        # Set up headers in row 7 with coloring
        headers_row_7 = [
            "HIOS Plan ID*\n(Standard Component)", "Plan Marketing Name*", "HIOS Product ID*",
            "Network ID*", "Service Area ID*", "Formulary ID*", "New/Existing Plan?*", "Plan Type*",
            "Level of Coverage*", "Design Type*", "Unique Plan Design?*", "QHP/Non-QHP*",
            "Notice Required\nfor Pregnancy*", "Plan Level Exclusions",
            "Limited Cost Sharing Plan Variation - Est Advanced Payment",
            "Does this plan offer Composite Rating?*", "Child-Only Offering*", "Child Only Plan ID",
            "Tobacco Wellness Program Offered*", "Disease Management Programs Offered",
            "EHB Percent of Total Premium", "EHB Apportionment for Pediatric Dental",
            "Guaranteed Rate", "Plan Effective Date*", "Plan Expiration Date",
            "Out of Country Coverage*", "Out of Country Coverage Description",
            "Out of Service Area Coverage*", "Out of Service Area Coverage Description",
            "National Network*"
        ]

        # Headers for row 20
        headers_row_20 = [
            "Out of Country Coverage Description",
            "Out of Service Area Coverage*", 
            "Out of Service Area Coverage Description",
            "National Network*", 
            "Benefits", 
            "EHB Is this Benefit Covered?",
            "Quantitative Limit on Service", 
            "Limit Quantity", 
            "Limit Unit", 
            "Exclusions", 
            "Benefit Explanation", 
            "Explanation",
        ]

        # Function to set headers
        def set_headers(ws, row, headers):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.fill = header_fill

        # Apply headers on row 7
        set_headers(ws, 7, headers_row_7)
        # Set column widths for visibility
        column_widths = [30, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 30, 30, 30, 30]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width

        # Populate filtered_left_join data starting from row 8
        def populate_data_from_df(ws, start_row, df):
            for index, row in df.iterrows():
                for col_index, (key, value) in enumerate(row.items()):
                    cell = ws.cell(row=start_row + index, column=col_index + 1)
                    cell.value = value
                    cell.border = thin_border

        populate_data_from_df(ws, start_row=8, df=filtered_left_join)
        # Apply headers on row 20
        set_headers(ws, 21, headers_row_20)

        # Set column widths for visibility
        column_widths = [30, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 30, 30, 30, 30]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width

        # Select the first 20 rows from BenefitInformation
        benefit_data = BenefitInformation.head(20)

        # Function to populate data from DataFrame
        def populate_data_from_df(ws, start_row, df):
            for index, row in df.iterrows():
                for col_index, (key, value) in enumerate(row.items()):
                    cell = ws.cell(row=start_row + index, column=col_index + 1)
                    cell.value = value
                    cell.border = thin_border

        populate_data_from_df(ws, start_row=22, df=benefit_data)  # Assuming data starts at row 21 below the headers in row 20

    # Create the Cost Share Variances sheet and apply formatting
    ws2 = wb.create_sheet(title="Cost Share Variances")

    # Set initial instructional content with colors
    instructions = [
        ("All Fields with an asterisk(*) are required", "A1:C1"),
        ("", "A2:A2"),
    ]
    # Function to apply instructions
    def apply_instructions(sheet):
        for text, cell_range in instructions:
            sheet.merge_cells(cell_range)
            cell = sheet[cell_range.split(':')[0]]  # Get the first cell of the range
            cell.value = text
            cell.alignment = Alignment(horizontal='center' if "A1:C1" in cell_range else 'left', vertical='center', wrap_text=True)
            # Apply light tan fill for instructional content
            cell.fill = instruction_fill

    apply_instructions(ws2)
    headers_sheet2  = [
                            "HIOS Plan ID (Standard Component + Variant)",
                            "Plan Variant Marketing Name*",
                            "Level of Coverage (Metal Level)",
                            "CSR Variation Type",
                            "Issuer Actuarial Value",
                            "AV Calculator Output Number*",
                            "Medical & Drug Deductibles Integrated?*",
                            "Medical & Drug Maximum Out of Pocket Integrated?*",
                            "Is a Referral Required for Specialist?*",
                            "Specialist(s) Requiring a Referral",
                            "Multiple In Network Tiers?*",
                            "1st Tier Utilization*",
                            "2nd Tier Utilization",
                            "Deductible*",
                            "Copayment*",
                            "Coinsurance*",
                            "Limit*",
                            "Deductible*",
                            "Copayment*",
                            "Coinsurance*",
                            "Limit*",
                            "Deductible*",
                            "Copayment*",
                            "Coinsurance*",
                            "Limit*",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Default Coinsurance",
                            "Individual",
                            "Family",
                            "Default Coinsurance",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Default Coinsurance",
                            "Individual",
                            "Family",
                            "Default Coinsurance",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "Default Coinsurance",
                            "Individual",
                            "Family",
                            "Default Coinsurance",
                            "Individual",
                            "Family",
                            "Individual",
                            "Family",
                            "HSA Eligible*",
                            "HSA/HRA Employer Contribution",
                            "HSA/HRA Employer Contribution Amount",
                            "Maximum Coinsurance for Specialty Drugs",
                            "Maximum Number of Days for Charging an Inpatient Copay?",
                            "Begin Primary Care Cost-Sharing After a Set Number of Visits?",
                            "Begin Primary Care Deductible/Coinsurance After a Set Number of Copays?",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network",
                            "In Network (Tier 1)",
                            "In Network (Tier 2)",
                            "Out of Network"
                        ]


    # Apply headers to Sheet2 starting from the third row
    set_headers(ws2, 3, headers_sheet2)

    # Add dummy data under headers in Sheet 2
    dummy_data = [
        ["Data 1-1", "Data 1-2", "Data 1-3", "Data 1-4", "Data 1-5"],
        ["Data 2-1", "Data 2-2", "Data 2-3", "Data 2-4", "Data 2-5"],
        ["Data 3-1", "Data 3-2", "Data 3-3", "Data 3-4", "Data 3-5"],
        ["Data 4-1", "Data 4-2", "Data 4-3", "Data 4-4", "Data 4-5"]
    ]

    # Populate dummy data starting from row 4
    for row_index, row_data in enumerate(dummy_data, start=4):
        for col_index, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_index, column=col_index)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = instruction_fill 
            cell.border = thin_border # Using the same light tan color for consistency



# Dummy data to match the headers list length
dummy_data = [["Data"] * len(headers_sheet2)] * 10  # Adjust the number of rows as needed

# Thin border and fill for the cells
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Populate dummy data starting from row 4
for row_index, row_data in enumerate(dummy_data, start=4):
    for col_index, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_index, column=col_index)
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = instruction_fill

    # Remove empty rows and sheets possibly created in a previous loop
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        max_row = ws.max_row
        for row in range(max_row, 0, -1):
            if all(ws.cell(row=row, column=col).value is None for col in range(1, ws.max_column + 1)):
                ws.delete_rows(row)

# Save the workbook
wb.save('Custom_Structured_Excel_File_v2.xlsx')
