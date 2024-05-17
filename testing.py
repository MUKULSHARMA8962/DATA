import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Set the title of the sheet and freeze panes
ws.title = "2025 Plans & Benefits Template v14.0"
ws.freeze_panes = 'A8'  # Freeze the rows above A8

# Define the borders and fill colors
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
# header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Gray color for headers
# instruction_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color for instructions
# Green light color for headers
header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Green light color for instructions
instruction_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
# Set initial instructional content with colors
instructions = [
    ("2025 Plans & Benefits Template v14.0", "A1:E1"),
    ("HIOS Issuer ID*: 32753", "A2:A2"),
    ("Issuer State*: MO", "A3:A3"),
    ("Market Coverage*: Individual", "A4:A4"),
    ("Dental Only Plan*: No", "A5:A5"),
    ("To use this template, please review the user guide and instructions. All fields with an asterisk (*) are required", "G1:K1"),
    ("you will need to save the latest version of the add-in file (PlansBenefitsAddInPY25.xlam) on your machine.", "G2:K2"),
    ("To create the cost share variance worksheet and enter the cost sharing amounts for both individual and SHOP (small group) markets, use the Create Cost Share Variances macro.", "G3:K3"),
    ("To create additional Benefits Package worksheets, use the Create New Benefits Package macro.", "G4:K4"),
    ("o populate the benefits on the Benefits Package worksheet with your State EHB Standards, use the Refresh EHB macro.", "G5:K5")
]     


for text, cell_range in instructions:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(':')[0]] 
    cell.value = text
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if "No" in text:  # Check if the text contains "No"
        cell.fill = instruction_fill
# Set up headers in row 7 with coloring
headers = [
    ("HIOS Plan ID*\n(Standard Component)", "A7"),
    ("Plan Marketing Name*", "B7"),
    ("HIOS Product ID*", "C7"),
    ("Network ID*", "D7"),
    ("Service Area ID*", "E7"),
    ("Formulary ID*", "F7"),
    ("New/Existing Plan?*", "G7"),
    ("Plan Type*", "H7"),
    ("Level of Coverage*", "I7"),
    ("Design Type*", "J7"),
    ("Unique Plan Design?*", "K7"),
    ("QHP/Non-QHP*", "L7"),
    ("Notice Required\nfor Pregnancy*", "M7"),
    ("Plan Level Exclusions", "N7"),
    ("Limited Cost Sharing Plan Variation - Est Advanced Payment", "O7"),
    ("Does this plan offer Composite Rating?*", "P7"),
    ("Child-Only Offering*", "Q7"),
    ("Child Only Plan ID", "R7"),
    ("Tobacco Wellness Program Offered*", "S7"),
    ("Disease Management Programs Offered", "T7"),
    ("EHB Percent of Total Premium", "U7"),
    ("EHB Apportionment for Pediatric Dental", "V7"),
    ("Guaranteed Rate", "W7"),
    ("Plan Effective Date*", "X7"),
    ("Plan Expiration Date", "Y7"),
    ("Out of Country Coverage*", "Z7"),
    ("Out of Country Coverage Description", "AA7"),
    ("Out of Service Area Coverage*", "AB7"),
    ("Out of Service Area Coverage Description", "AC7"),
    ("National Network*", "AD7"),
     ("Benefits", "A20") ,
     ("EHB	Is this Benefit Covered?", "B20") ,
     ("Quantitative Limit on Service", "C20") ,
     ("Limit Quantity", "D20") ,
     ("Limit Unit", "E20") ,
     ("Exclusions", "F20") ,
     ("Benefit Explanation", "G20") ,
     ("Explanation", "H20") ,
  

]

# Apply headers and format cells with coloring
for header, cell_reference in headers:
    cell = ws[cell_reference]
    cell.value = header
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.fill = header_fill  # Applying gray fill for header cells

# Set column widths for visibility
column_widths = [30, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 30, 30, 30, 30]
for i, column_width in enumerate(column_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width







# Create Sheet 2 and apply formatting
ws2 = wb.create_sheet(title="Sheet2")
# Example formatting for Sheet 2
# You can customize this based on your requirements
for row in ws2.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5):
    for cell in row:
        cell.value = "Sheet 2 Cell"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = instruction_fill  # Applying yellow fill color for instructional content

# Save the workbook
wb.save('Custom_Structured_Excel_File_v2.xlsx')
# Save the workbook
wb.save('Custom_Structured_Excel_File_v2.xlsx')
